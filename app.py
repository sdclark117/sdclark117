import json
import logging
import os
import secrets
import sys
import time
from datetime import datetime, timedelta
from functools import wraps
from io import BytesIO
from typing import Any, Dict, Optional

import click
import googlemaps
import gspread
import pandas as pd
import stripe
from dotenv import load_dotenv
from flask import (
    Flask,
    flash,
    jsonify,
    redirect,
    render_template,
    request,
    send_file,
    session,
    url_for,
)
from flask.cli import with_appcontext
from flask_login import (
    LoginManager,
    UserMixin,
    current_user,
    login_required,
    login_user,
    logout_user,
)
from flask_mail import Mail, Message
from flask_migrate import Migrate
from flask_sqlalchemy import SQLAlchemy
from google.oauth2.service_account import Credentials
from openpyxl import Workbook
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename

# Load environment variables
load_dotenv()

app = Flask(__name__)
secret_key = os.getenv("SECRET_KEY") or os.getenv("FLASK_SECRET_KEY")
if not secret_key:
    # Generate a secure random key if none is provided
    secret_key = secrets.token_hex(32)
    print(
        "WARNING: No SECRET_KEY or FLASK_SECRET_KEY environment variable set. Generated a temporary key."
    )
    print(
        "Please set SECRET_KEY or FLASK_SECRET_KEY environment variable for production use."
    )
app.secret_key = secret_key

# Configure Flask for production - only set SERVER_NAME if explicitly provided
if os.getenv("SERVER_NAME"):
    app.config["SERVER_NAME"] = os.getenv("SERVER_NAME")
app.config["PREFERRED_URL_SCHEME"] = (
    "https" if os.getenv("FLASK_ENV") == "production" else "http"
)

# Logging setup
app.logger.setLevel(logging.INFO)
handler = logging.StreamHandler(sys.stdout)
handler.setLevel(logging.INFO)
if not any(isinstance(h, logging.StreamHandler) for h in app.logger.handlers):
    app.logger.addHandler(handler)

# Gunicorn integration
if "gunicorn" in os.environ.get("SERVER_SOFTWARE", ""):
    gunicorn_logger = logging.getLogger("gunicorn.error")
    app.logger.handlers = gunicorn_logger.handlers
    app.logger.setLevel(gunicorn_logger.level)

# Database Configuration
db_url = os.environ.get("DATABASE_URL")
if db_url and db_url.startswith("postgres://"):
    db_url = db_url.replace("postgres://", "postgresql://", 1)
elif not db_url:
    # Use persistent SQLite database file
    db_url = "sqlite:///sneaker_agent.db"
    print(f"📁 Using persistent database: {db_url}")

app.config["SQLALCHEMY_DATABASE_URI"] = db_url
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

app.config["REMEMBER_COOKIE_DURATION"] = timedelta(days=30)
app.config["REMEMBER_COOKIE_HTTPONLY"] = True
app.config["REMEMBER_COOKIE_SECURE"] = os.environ.get("FLASK_ENV") == "production"
app.config["GOOGLE_API_KEY"] = os.environ.get("GOOGLE_MAPS_API_KEY")

# File upload configuration
app.config["UPLOAD_FOLDER"] = "static/profile_pictures"
app.config["MAX_CONTENT_LENGTH"] = 16 * 1024 * 1024  # 16MB max file size
ALLOWED_EXTENSIONS = {"png", "jpg", "jpeg", "gif"}

# Create upload folder if it doesn't exist
os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)

# Email configuration
app.config["MAIL_SERVER"] = "smtp.gmail.com"
app.config["MAIL_PORT"] = 587
app.config["MAIL_USE_TLS"] = True
app.config["MAIL_USERNAME"] = os.getenv("GMAIL_USERNAME2")
app.config["MAIL_PASSWORD"] = os.getenv("GMAIL_APP_PASSWORD2")
app.config["MAIL_DEFAULT_SENDER"] = os.getenv("GMAIL_USERNAME2")

# Debug email configuration
app.logger.info(f"📧 FLASK_ENV: {os.getenv('FLASK_ENV')}")
app.logger.info(f"📧 GMAIL_USERNAME2: {os.getenv('GMAIL_USERNAME2')}")
app.logger.info(f"📧 GMAIL_APP_PASSWORD2 set: {bool(os.getenv('GMAIL_APP_PASSWORD2'))}")
app.logger.info(f"📧 GMAIL_USERNAME set: {bool(app.config['MAIL_USERNAME'])}")
app.logger.info(f"📧 GMAIL_APP_PASSWORD set: {bool(app.config['MAIL_PASSWORD'])}")
app.logger.info(f"📧 MAIL_SERVER: {app.config['MAIL_SERVER']}")
app.logger.info(f"📧 MAIL_USERNAME value: {app.config['MAIL_USERNAME']}")
app.logger.info(
    f"📧 MAIL_PASSWORD length: {len(app.config['MAIL_PASSWORD']) if app.config['MAIL_PASSWORD'] else 0}"
)

# Check if Gmail is configured
gmail_configured = app.config["MAIL_USERNAME"] and app.config["MAIL_PASSWORD"]

if not gmail_configured:
    if os.getenv("FLASK_ENV") == "production":
        app.logger.error("❌ Gmail not configured in PRODUCTION mode!")
        app.logger.error(
            "❌ Please set GMAIL_USERNAME2 and GMAIL_APP_PASSWORD2 in Render"
        )
    else:
        app.logger.warning(
            "📧 Gmail not configured. Using email simulation for development."
        )
        # For development, we'll simulate email sending
        app.config["MAIL_SERVER"] = "localhost"
        app.config["MAIL_PORT"] = 1025
        app.config["MAIL_USE_TLS"] = False
        app.config["MAIL_USERNAME"] = "noreply@businessleadfinder.com"
        app.config["MAIL_PASSWORD"] = ""  # nosec B105
        app.config["MAIL_DEFAULT_SENDER"] = "noreply@businessleadfinder.com"
        app.logger.info("📧 Using development email simulation")
else:
    app.logger.info("✅ Email configuration loaded successfully.")
    app.logger.info(f"📧 Using Gmail SMTP: {app.config['MAIL_USERNAME']}")

mail = Mail(app)
db = SQLAlchemy(app)
migrate = Migrate(app, db)


@click.command("init-db")
@with_appcontext
def init_db_command():
    """Clear the existing data and create new tables."""
    db.drop_all()
    db.create_all()
    click.echo("Initialized the database.")


app.cli.add_command(init_db_command)

# Initialize Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = "index"  # type: ignore


@login_manager.unauthorized_handler
def unauthorized_callback():
    """Handle unauthorized access attempts."""
    if request.path.startswith("/api/"):
        return jsonify(error="Authentication required to access this endpoint."), 401
    flash("You must be logged in to view this page.")
    return redirect(url_for("index"))


class User(db.Model, UserMixin):
    __tablename__ = "users"
    id = db.Column(db.Integer, primary_key=True)
    email = db.Column(db.String(120), unique=True, nullable=False)
    password_hash = db.Column(db.String(512), nullable=False)
    name = db.Column(db.String(100))
    business = db.Column(db.String(100))
    phone = db.Column(db.String(20))
    profile_picture = db.Column(db.String(255))  # Store file path
    is_verified = db.Column(db.Boolean, default=False, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    trial_ends_at = db.Column(db.DateTime)
    is_admin = db.Column(db.Boolean, default=False, nullable=False)
    is_support = db.Column(db.Boolean, default=False, nullable=False)
    is_technical = db.Column(db.Boolean, default=False, nullable=False)
    staff_role = db.Column(db.String(50))  # 'support', 'technical', etc.
    stripe_customer_id = db.Column(db.String(120), unique=True)
    stripe_subscription_id = db.Column(db.String(120), unique=True)
    current_plan = db.Column(db.String(50))
    search_count = db.Column(db.Integer, default=0)
    last_search_reset = db.Column(db.DateTime)
    settings = db.relationship(
        "UserSettings", backref="user", uselist=False, cascade="all, delete-orphan"
    )

    def __init__(self, **kwargs):
        super(User, self).__init__(**kwargs)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class EmailVerificationToken(db.Model):
    __tablename__ = "email_verification_tokens"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    token = db.Column(db.String(128), unique=True, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __init__(self, user_id, token, expires_at):
        self.user_id = user_id
        self.token = token
        self.expires_at = expires_at


class PasswordResetToken(db.Model):
    __tablename__ = "password_reset_tokens"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    token = db.Column(db.String(128), unique=True, nullable=False)
    expires_at = db.Column(db.DateTime, nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)

    def __init__(self, user_id, token, expires_at):
        self.user_id = user_id
        self.token = token
        self.expires_at = expires_at


class GuestUsage(db.Model):
    __tablename__ = "guest_usage"
    id = db.Column(db.Integer, primary_key=True)
    ip_address = db.Column(db.String(45), nullable=False, index=True)  # IPv6 compatible
    user_agent = db.Column(db.String(500))
    search_count = db.Column(db.Integer, default=0)
    first_visit = db.Column(db.DateTime, default=datetime.utcnow)
    last_visit = db.Column(db.DateTime, default=datetime.utcnow)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(
        db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow
    )

    def __init__(self, ip_address, user_agent):
        self.ip_address = ip_address
        self.user_agent = user_agent


class UserSettings(db.Model):
    __tablename__ = "user_settings"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(
        db.Integer, db.ForeignKey("users.id"), unique=True, nullable=False
    )
    default_radius = db.Column(db.Integer, default=5)
    default_business_type = db.Column(db.String(100))
    remember_last_search = db.Column(db.Boolean, default=False)
    results_per_page = db.Column(db.Integer, default=25)
    show_map_by_default = db.Column(db.Boolean, default=True)
    email_notifications = db.Column(db.Boolean, default=True)
    search_reminders = db.Column(db.Boolean, default=False)
    last_search_city = db.Column(db.String(100))
    last_search_state = db.Column(db.String(100))
    last_search_business_type = db.Column(db.String(100))
    last_search_radius = db.Column(db.Integer)
    updated_at = db.Column(
        db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow
    )


class SiteAnalytics(db.Model):
    __tablename__ = "site_analytics"
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False, unique=True)
    total_visits = db.Column(db.Integer, default=0)
    unique_visitors = db.Column(db.Integer, default=0)
    registered_users = db.Column(db.Integer, default=0)
    active_users = db.Column(db.Integer, default=0)
    searches_performed = db.Column(db.Integer, default=0)
    exports_performed = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(
        db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow
    )


class PageVisits(db.Model):
    __tablename__ = "page_visits"
    id = db.Column(db.Integer, primary_key=True)
    date = db.Column(db.Date, nullable=False)
    page = db.Column(db.String(100), nullable=False)
    visits = db.Column(db.Integer, default=0)
    unique_visitors = db.Column(db.Integer, default=0)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    updated_at = db.Column(
        db.DateTime, default=datetime.utcnow, onupdate=datetime.utcnow
    )

    __table_args__ = (db.UniqueConstraint("date", "page"),)


class UserActivity(db.Model):
    __tablename__ = "user_activity"
    id = db.Column(db.Integer, primary_key=True)
    user_id = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=True)
    ip_address = db.Column(db.String(45), nullable=False)
    action = db.Column(db.String(100), nullable=False)
    page = db.Column(db.String(100), nullable=True)
    user_agent = db.Column(db.String(500), nullable=True)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class AITeam(db.Model):
    __tablename__ = "ai_teams"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    manager_id = db.Column(db.Integer, db.ForeignKey("ai_managers.id"))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    agents = db.relationship("AIAgent", backref="team", cascade="all, delete-orphan")
    manager = db.relationship("AIManager", back_populates="team", uselist=False)


class AIManager(db.Model):
    __tablename__ = "ai_managers"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    team = db.relationship("AITeam", back_populates="manager", uselist=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)


class AIAgent(db.Model):
    __tablename__ = "ai_agents"
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), nullable=False)
    specialty = db.Column(db.String(100), nullable=False)
    team_id = db.Column(db.Integer, db.ForeignKey("ai_teams.id"))
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    status = db.Column(db.String(20), default="Active")


class StaffAccessCode(db.Model):
    __tablename__ = "staff_access_codes"
    id = db.Column(db.Integer, primary_key=True)
    code = db.Column(db.String(8), unique=True, nullable=False)
    staff_role = db.Column(db.String(50), nullable=False)
    is_support = db.Column(db.Boolean, default=False)
    is_technical = db.Column(db.Boolean, default=False)
    created_by = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=False)
    created_at = db.Column(db.DateTime, default=datetime.utcnow)
    expires_at = db.Column(db.DateTime, nullable=False)
    is_used = db.Column(db.Boolean, default=False)
    used_by = db.Column(db.Integer, db.ForeignKey("users.id"), nullable=True)
    used_at = db.Column(db.DateTime, nullable=True)

    def __init__(self, **kwargs):
        super().__init__(**kwargs)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def generate_token():
    return secrets.token_urlsafe(32)


def allowed_file(filename):
    """Check if file extension is allowed."""
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


def save_profile_picture(file, user_id):
    """Save profile picture and return the filename."""
    if file and allowed_file(file.filename):
        # Create secure filename
        filename = secure_filename(file.filename)
        # Add user ID to make filename unique
        name, ext = filename.rsplit(".", 1)
        filename = f"user_{user_id}_{int(time.time())}.{ext}"

        # Save file
        filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
        file.save(filepath)

        return filename
    return None


def get_client_ip():
    """Get the client's IP address, handling proxies."""
    # Check for forwarded headers first (common with proxies)
    if request.headers.get("X-Forwarded-For"):
        return request.headers.get("X-Forwarded-For").split(",")[0].strip()
    elif request.headers.get("X-Real-IP"):
        return request.headers.get("X-Real-IP")
    elif request.headers.get("X-Client-IP"):
        return request.headers.get("X-Client-IP")
    else:
        return request.remote_addr


def get_or_create_guest_usage():
    """Get or create guest usage tracking for the current IP."""
    if current_user.is_authenticated:
        return None

    ip_address = get_client_ip()
    user_agent = request.headers.get("User-Agent", "")

    # Try to find existing guest usage
    guest_usage = GuestUsage.query.filter_by(ip_address=ip_address).first()

    if guest_usage:
        # Update last visit
        guest_usage.last_visit = datetime.utcnow()
        db.session.commit()
        return guest_usage
    else:
        # Create new guest usage record
        guest_usage = GuestUsage(ip_address=ip_address, user_agent=user_agent)
        db.session.add(guest_usage)
        db.session.commit()
        return guest_usage


def track_page_visit(page_name):
    """Track a page visit for analytics."""
    try:
        today = datetime.utcnow().date()
        ip_address = get_client_ip()
        user_agent = request.headers.get("User-Agent", "")

        # Get or create page visit record for today
        page_visit = PageVisits.query.filter_by(date=today, page=page_name).first()
        if not page_visit:
            page_visit = PageVisits(date=today, page=page_name)
            db.session.add(page_visit)

        # Increment total visits
        page_visit.visits = (page_visit.visits or 0) + 1

        # Check if this is a unique visitor (by IP)
        existing_activity = UserActivity.query.filter(
            db.func.date(UserActivity.created_at) == today,
            UserActivity.ip_address == ip_address,
            UserActivity.page == page_name,
        ).first()

        if not existing_activity:
            page_visit.unique_visitors = (page_visit.unique_visitors or 0) + 1

        # Record user activity
        activity = UserActivity(
            user_id=current_user.id if current_user.is_authenticated else None,
            ip_address=ip_address,
            action="page_visit",
            page=page_name,
            user_agent=user_agent,
        )
        db.session.add(activity)

        db.session.commit()
    except Exception as e:
        app.logger.error(f"Error tracking page visit: {e}")
        db.session.rollback()


def track_user_action(action, page_name=None):
    """Track a user action for analytics."""
    try:
        ip_address = get_client_ip()
        user_agent = request.headers.get("User-Agent", "")

        activity = UserActivity(
            user_id=current_user.id if current_user.is_authenticated else None,
            ip_address=ip_address,
            action=action,
            page=page_name,
            user_agent=user_agent,
        )
        db.session.add(activity)
        db.session.commit()
    except Exception as e:
        app.logger.error(f"Error tracking user action: {e}")
        db.session.rollback()


def update_daily_analytics():
    """Update daily analytics summary."""
    try:
        today = datetime.utcnow().date()

        # Get or create analytics record for today
        analytics = SiteAnalytics.query.filter_by(date=today).first()
        if not analytics:
            analytics = SiteAnalytics(date=today)
            db.session.add(analytics)

        # Count total visits today
        total_visits = UserActivity.query.filter(
            db.func.date(UserActivity.created_at) == today
        ).count()

        # Count unique visitors today
        unique_visitors = (
            db.session.query(db.func.count(db.func.distinct(UserActivity.ip_address)))
            .filter(db.func.date(UserActivity.created_at) == today)
            .scalar()
        )

        # Count registered users
        registered_users = User.query.count()

        # Count active users today (users who performed actions)
        active_users = (
            db.session.query(db.func.count(db.func.distinct(UserActivity.user_id)))
            .filter(
                db.func.date(UserActivity.created_at) == today,
                UserActivity.user_id.isnot(None),
            )
            .scalar()
        )

        # Count searches performed today
        searches_performed = UserActivity.query.filter(
            db.func.date(UserActivity.created_at) == today,
            UserActivity.action == "search",
        ).count()

        # Count exports performed today
        exports_performed = UserActivity.query.filter(
            db.func.date(UserActivity.created_at) == today,
            UserActivity.action == "export",
        ).count()

        # Update analytics
        analytics.total_visits = total_visits
        analytics.unique_visitors = unique_visitors or 0
        analytics.registered_users = registered_users
        analytics.active_users = active_users or 0
        analytics.searches_performed = searches_performed
        analytics.exports_performed = exports_performed

        db.session.commit()
    except Exception as e:
        app.logger.error(f"Error updating daily analytics: {e}")
        db.session.rollback()


def get_analytics_data(days=30):
    """Get analytics data for the specified number of days."""
    try:
        end_date = datetime.utcnow().date()
        start_date = end_date - timedelta(days=days)

        # Get site analytics
        site_analytics = (
            SiteAnalytics.query.filter(
                SiteAnalytics.date >= start_date, SiteAnalytics.date <= end_date
            )
            .order_by(SiteAnalytics.date)
            .all()
        )

        # Get page visits
        page_visits = (
            PageVisits.query.filter(
                PageVisits.date >= start_date, PageVisits.date <= end_date
            )
            .order_by(PageVisits.date)
            .all()
        )

        # Get user activity summary
        user_activity = (
            db.session.query(
                UserActivity.action, db.func.count(UserActivity.id).label("count")
            )
            .filter(
                db.func.date(UserActivity.created_at) >= start_date,
                db.func.date(UserActivity.created_at) <= end_date,
            )
            .group_by(UserActivity.action)
            .all()
        )

        return {
            "site_analytics": site_analytics,
            "page_visits": page_visits,
            "user_activity": user_activity,
        }
    except Exception as e:
        app.logger.error(f"Error getting analytics data: {e}")
        return {"site_analytics": [], "page_visits": [], "user_activity": []}


def send_email(subject, recipients, body, html_body=None):
    try:
        # Check if we're in development mode and Gmail is not configured
        if app.config.get("MAIL_SERVER") == "localhost":
            # Simulate email sending for development
            app.logger.info(f"📧 [DEV MODE] Email would be sent to {recipients}")
            app.logger.info(f"📧 Subject: {subject}")
            app.logger.info(f"📧 Body: {body[:100]}...")

            # In development, we'll simulate successful email sending
            # In production, this would be replaced with actual email service
            return True

        # Check if Gmail credentials are configured
        if not app.config.get("MAIL_USERNAME") or not app.config.get("MAIL_PASSWORD"):
            if os.getenv("FLASK_ENV") == "production":
                app.logger.error(
                    "❌ Gmail credentials not configured in PRODUCTION mode!"
                )
                app.logger.error(
                    "❌ Please set GMAIL_USERNAME2 and GMAIL_APP_PASSWORD2 in Render"
                )
            else:
                app.logger.error(
                    "Gmail credentials not configured. Please set GMAIL_USERNAME2 and "
                    "GMAIL_APP_PASSWORD2 environment variables."
                )
            return False

        # Create message
        msg = Message(subject, recipients=recipients)
        msg.body = body
        if html_body:
            msg.html = html_body

        # Send email
        mail.send(msg)
        app.logger.info(f"✅ Email sent successfully to {recipients}")
        return True
    except Exception as e:
        app.logger.error(f"❌ Email sending error: {e}")
        return False


def send_verification_email(user_id, email, name):
    try:
        token = generate_token()
        expires_at = datetime.utcnow() + timedelta(hours=24)
        new_token = EmailVerificationToken(
            user_id=user_id, token=token, expires_at=expires_at
        )
        db.session.add(new_token)
        db.session.commit()

        verification_url = url_for("verify_email", token=token, _external=True)
        subject = "Verify Your Email - Business Lead Finder"
        html_body = render_template(
            "email_verification.html", name=name, verification_url=verification_url
        )
        body = (
            f"Hello {name or 'there'}! Please verify your email by clicking this link: "
            f"{verification_url}"
        )

        success = send_email(subject, [email], body, html_body)

        # In development mode, show the verification link in console
        if app.config.get("MAIL_SERVER") == "localhost":
            print(f"\n🔗 [DEV MODE] Email verification link for {email}:")
            print(f"🔗 {verification_url}")
            print(f"🔗 Token: {token}")
            print(f"🔗 Expires: {expires_at}\n")

        if success:
            app.logger.info(f"Verification email sent successfully to {email}")
        else:
            app.logger.error(f"Failed to send verification email to {email}")
        return success
    except Exception as e:
        app.logger.error(f"Error in send_verification_email: {e}")
        return False


def send_password_reset_email(user_id, email, name):
    token = generate_token()
    expires_at = datetime.utcnow() + timedelta(hours=1)
    new_token = PasswordResetToken(user_id=user_id, token=token, expires_at=expires_at)
    db.session.add(new_token)
    db.session.commit()

    reset_url = url_for("reset_password_page", token=token, _external=True)
    subject = "Password Reset - Business Lead Finder"
    html_body = render_template("password_reset.html", name=name, reset_url=reset_url)
    body = (
        f"Hello {name or 'there'}! Please reset your password by clicking this link: "
        f"{reset_url}"
    )

    return send_email(subject, [email], body, html_body)


def cleanup_expired_tokens():
    """Clean up expired verification and reset tokens."""
    try:
        now = datetime.utcnow()

        # Clean up expired verification tokens
        expired_verification_tokens = (
            db.session.query(EmailVerificationToken)
            .filter(EmailVerificationToken.expires_at < now)
            .all()
        )  # type: ignore
        for token in expired_verification_tokens:
            db.session.delete(token)

        # Clean up expired reset tokens
        expired_reset_tokens = (
            db.session.query(PasswordResetToken)
            .filter(PasswordResetToken.expires_at < now)
            .all()
        )  # type: ignore
        for token in expired_reset_tokens:
            db.session.delete(token)

        db.session.commit()
        app.logger.info(
            f"Cleaned up {len(expired_verification_tokens)} expired verification tokens "
            f"and {len(expired_reset_tokens)} expired reset tokens"
        )
    except Exception as e:
        app.logger.error(f"Error cleaning up expired tokens: {e}")
        db.session.rollback()


def reset_guest_usage_daily():
    """Reset guest usage counts daily to allow new free searches."""
    try:
        # Reset search counts for guest usage older than 24 hours
        yesterday = datetime.utcnow() - timedelta(days=1)

        guest_usage_to_reset = GuestUsage.query.filter(
            GuestUsage.updated_at < yesterday
        ).all()

        for guest_usage in guest_usage_to_reset:
            guest_usage.search_count = 0
            guest_usage.updated_at = datetime.utcnow()

        db.session.commit()

        if guest_usage_to_reset:
            app.logger.info(
                f"Reset search counts for {len(guest_usage_to_reset)} guest IPs"
            )

    except Exception as e:
        app.logger.error(f"Error resetting guest usage: {e}")
        db.session.rollback()


def get_coordinates(location_query: str, api_key: str) -> Optional[Dict[str, float]]:
    """Get coordinates for a location query using Google Maps Geocoding API."""
    if not api_key:
        app.logger.error("No Google API key provided for geocoding")
        return None

    # Log API key info (without exposing the full key)
    api_key_preview = api_key[:10] + "..." if len(api_key) > 10 else api_key
    app.logger.info(f"Using API key: {api_key_preview}")

    gmaps = googlemaps.Client(key=api_key)
    try:
        app.logger.info(f"Geocoding location query: '{location_query}'")
        geocode_result = gmaps.geocode(location_query)  # type: ignore

        # Log the raw response for debugging
        app.logger.info(f"Geocoding raw response: {geocode_result}")

        if geocode_result:
            location = geocode_result[0]["geometry"]["location"]
            app.logger.info(f"Geocoding result for '{location_query}': {location}")
            return {"lat": location["lat"], "lng": location["lng"]}
        else:
            app.logger.warning(
                f"Geocoding returned no results for query: '{location_query}'"
            )
            return None
    except googlemaps.exceptions.ApiError as e:
        app.logger.error(f"Geocoding API error for query '{location_query}': {e}")
        app.logger.error(f"API Error details: {str(e)}")
        return None
    except Exception as e:
        app.logger.error(
            f"An unexpected error occurred during geocoding for query "
            f"'{location_query}': {e}"
        )
        return None


# Add this near the top of app.py, after imports
CATEGORY_FALLBACKS = {
    "food": "restaurant",
    "shop": "store",
    "shopping": "store",
    "drinks": "bar",
    "coffee": "cafe",
    "eat": "restaurant",
    "groceries": "supermarket",
    "market": "supermarket",
    "gas": "gas_station",
    "gas station": "gas_station",
    "pharmacy": "pharmacy",
    "doctor": "doctor",
    "health": "hospital",
    "fitness": "gym",
    "workout": "gym",
    "hotel": "lodging",
    "stay": "lodging",
    "car": "car_repair",
    "auto": "car_repair",
    "mechanic": "car_repair",
    # Add more as needed
}


def search_places(lat, lng, business_type, radius, api_key, max_reviews=100):
    """Search for places using Google Places API."""
    if not api_key:
        app.logger.error("No Google API key provided for places search")
        return [], {"lat": lat, "lng": lng}

    gmaps = googlemaps.Client(key=api_key)
    all_leads = []
    seen_place_ids = set()

    # Fallback to a broader category if needed
    search_term = CATEGORY_FALLBACKS.get(
        str(business_type).strip().lower(), business_type
    )
    app.logger.info(
        f"Searching places: lat={lat}, lng={lng}, radius={radius}, "
        f"business_type='{business_type}', search_term='{search_term}', "
        f"max_reviews={max_reviews}"
    )

    # Initial search request
    try:
        places_result = gmaps.places_nearby(  # type: ignore
            location=(lat, lng), radius=radius, keyword=search_term, language="en"
        )
    except Exception as e:
        app.logger.error(f"Error calling Google Places API: {e}")
        return [], {"lat": lat, "lng": lng}

    total_google_results = 0
    total_after_filter = 0

    while True:
        api_response_data = places_result
        if not api_response_data or api_response_data.get("status") != "OK":
            break
        current_results = api_response_data.get("results", [])
        total_google_results += len(current_results)
        for place in current_results:
            place_id = place.get("place_id")
            if place_id in seen_place_ids:
                continue
            seen_place_ids.add(place_id)
            if is_potential_lead(place):
                try:
                    details = get_place_details(place_id, api_key)
                    if details:
                        user_ratings_total = details.get("user_ratings_total")
                        if max_reviews is not None and user_ratings_total is not None:
                            if user_ratings_total > max_reviews:
                                continue
                        lead_data = {
                            "place_id": place_id,
                            "name": details.get("name"),
                            "address": details.get("formatted_address"),
                            "lat": details["geometry"]["location"]["lat"],
                            "lng": details["geometry"]["location"]["lng"],
                            "rating": details.get("rating"),
                            "website": details.get("website"),
                            "phone": details.get("formatted_phone_number"),
                            "opening_hours": format_opening_hours(
                                details.get("opening_hours", {})
                            ),
                            "reviews": user_ratings_total,
                            "business_type": format_business_types(
                                details.get("types", [])
                            ),
                            "business_status": place.get("business_status"),
                        }
                        all_leads.append(lead_data)
                        total_after_filter += 1
                except Exception as e:
                    app.logger.error(
                        (
                            "Error processing place details for '"
                            + str(place.get("name"))[:40]
                            + str(place.get("name"))[40:]
                            + "': "
                            + str(e)[:40]
                            + str(e)[40:]
                        )
                    )
        next_page_token = api_response_data.get("next_page_token")
        if not next_page_token:
            break
        time.sleep(2)
        try:
            places_result = gmaps.places_nearby(page_token=next_page_token)  # type: ignore
        except Exception as e:
            app.logger.error(f"Error fetching next page from Google Places API: {e}")
            break
    final_leads = list({lead["place_id"]: lead for lead in all_leads}.values())
    app.logger.info(
        f"Total Google results: {total_google_results}, After filtering: "
        f"{total_after_filter}, Final unique leads: {len(final_leads)}"
    )
    return final_leads, {"lat": lat, "lng": lng}


def get_place_details(place_id: str, api_key: str) -> Optional[Dict[str, Any]]:
    """Get detailed information about a place using Google Places API."""
    gmaps = googlemaps.Client(key=api_key)
    fields = [
        "name",
        "formatted_address",
        "international_phone_number",
        "website",
        "rating",
        "user_ratings_total",
        "opening_hours",
        "geometry",
    ]
    try:
        place_details = gmaps.place(place_id=place_id, fields=fields)  # type: ignore
        if place_details and place_details.get("result"):
            return place_details["result"]
        else:
            app.logger.warning(f"Could not get details for place_id: {place_id}")
            return None
    except googlemaps.exceptions.ApiError as e:
        app.logger.error(
            (
                "API Error getting place details for "
                + str(place_id)[:40]
                + str(place_id)[40:]
                + ": "
                + str(e)[:40]
                + str(e)[40:]
            )
        )
        return None
    except Exception as e:
        app.logger.error(f"Unexpected error getting place details for {place_id}: {e}")
        return None


def is_potential_lead(place):
    """Check if a place is a potential lead based on business status."""
    return place.get("business_status") == "OPERATIONAL"


def format_opening_hours(hours_data):
    """Format opening hours data for display."""
    if not hours_data or not hours_data.get("weekday_text"):
        return "Not available"
    return "\n".join(hours_data["weekday_text"])


def format_business_types(types):
    """Format business types for display."""
    if not types:
        return "N/A"
    return ", ".join(t.replace("_", " ").title() for t in types)


@app.route("/")
def index():
    # Restore the use of the environment variable for the API key
    google_maps_api_key = os.environ.get("GOOGLE_MAPS_API_KEY")
    if not google_maps_api_key:
        app.logger.warning("❌ GOOGLE_MAPS_API_KEY not found in environment variables")
        flash("Google Maps API key not configured.", "warning")

    # Get or create guest usage tracking
    guest_usage = get_or_create_guest_usage()

    # Track page visit
    track_page_visit("index")

    return render_template(
        "index.html",
        google_maps_api_key=google_maps_api_key,
        guest_usage=guest_usage,
    )


@app.route("/verify-email/<token>")
def verify_email(token: str):
    """Verify user email with token."""
    try:
        verification_record = EmailVerificationToken.query.filter_by(
            token=token
        ).first()
        if verification_record and verification_record.expires_at > datetime.utcnow():
            user = User.query.get(verification_record.user_id)
            if user:
                user.is_verified = True
                db.session.delete(verification_record)
                db.session.commit()
                return render_template(
                    "email_verification_result.html",
                    success=True,
                    message="Your email has been verified successfully! You can now log in.",
                )
            else:
                return render_template(
                    "email_verification_result.html",
                    success=False,
                    message="User not found.",
                )
        else:
            return render_template(
                "email_verification_result.html",
                success=False,
                message="Invalid or expired verification link.",
            )
    except Exception as e:
        app.logger.error(f"Error during email verification: {e}")
        return render_template(
            "email_verification_result.html",
            success=False,
            message="An error occurred during verification.",
        )


@app.route("/reset-password/<token>")
def reset_password_page(token):
    """Display password reset page."""
    try:
        reset_record = PasswordResetToken.query.filter_by(token=token).first()
        if reset_record and reset_record.expires_at > datetime.utcnow():
            return render_template("password_reset.html", token=token)
        else:
            flash("Invalid or expired password reset link.", "danger")
    except Exception as e:
        app.logger.error(f"Error during password reset page access: {e}")
        flash("An error occurred.", "danger")
    return redirect(url_for("index"))


@app.route("/api/register", methods=["POST"])
def register():
    """Register a new user."""
    try:
        data = request.get_json()
        if not data or not all(k in data for k in ["email", "password", "name"]):
            return jsonify(error="Missing required data"), 400

        email = data["email"].lower().strip()
        if User.query.filter_by(email=email).first():
            return jsonify(error="Email already exists"), 409

        password_hash = generate_password_hash(data["password"])

        new_user = User(
            email=email,
            password_hash=password_hash,
            name=data["name"].strip(),
            trial_ends_at=datetime.utcnow() + timedelta(days=7),
        )

        db.session.add(new_user)
        db.session.commit()

        send_verification_email(new_user.id, new_user.email, new_user.name)
        login_user(new_user, remember=True)

        return jsonify(message="Registration successful!"), 201
    except Exception as e:
        app.logger.error(f"Error during registration: {e}")
        db.session.rollback()
        return jsonify(error="An error occurred during registration."), 500


@app.route("/api/login", methods=["POST"])
def login():
    """Log in a user."""
    try:
        data = request.get_json()
        app.logger.info(f"Login attempt - data received: {bool(data)}")

        if not data or not data.get("email") or not data.get("password"):
            app.logger.warning("Login failed - missing email or password")
            return jsonify(error="Missing email or password"), 400

        email = data["email"].lower().strip()
        password = data["password"]
        app.logger.info(f"Login attempt for email: {email}")

        user = User.query.filter_by(email=email).first()
        app.logger.info(f"User found: {bool(user)}")

        if user:
            password_check = user.check_password(password)
            app.logger.info(f"Password check result: {password_check}")

            if password_check:
                login_user(user, remember=data.get("remember", False))
                app.logger.info(f"Login successful for user: {user.email}")
                return jsonify(message="Login successful"), 200
            else:
                app.logger.warning(
                    f"Login failed - incorrect password for user: {email}"
                )
        else:
            app.logger.warning(f"Login failed - user not found: {email}")

        return jsonify(error="Invalid email or password"), 401
    except Exception as e:
        app.logger.error(f"Error during login: {e}")
        return jsonify(error="An error occurred during login."), 500


@app.route("/api/logout", methods=["POST"])
@login_required
def logout():
    """Log out the current user."""
    try:
        user_email = current_user.email  # Store email before logout
        app.logger.info(f"🔧 Logout attempt for user: {user_email}")
        logout_user()
        app.logger.info(f"✅ Logout successful for user: {user_email}")
        return jsonify(message="Logout successful"), 200
    except Exception as e:
        app.logger.error(f"❌ Logout error: {e}")
        return jsonify(error="Logout failed"), 500


@app.route("/api/profile", methods=["PUT"])
@login_required
def update_profile():
    """Update user profile information."""
    try:
        data = request.get_json()
        if not data:
            return jsonify(error="Invalid data"), 400

        current_user.name = data.get("name", current_user.name)
        current_user.business = data.get("business", current_user.business)
        current_user.phone = data.get("phone", current_user.phone)
        db.session.commit()

        return jsonify(message="Profile updated successfully"), 200
    except Exception as e:
        app.logger.error(f"Error updating profile: {e}")
        db.session.rollback()
        return jsonify(error="An error occurred while updating profile."), 500


@app.route("/api/profile-picture", methods=["POST"])
@login_required
def upload_profile_picture():
    """Upload profile picture."""
    try:
        if "profile_picture" not in request.files:
            return jsonify(error="No file provided"), 400

        file = request.files["profile_picture"]
        if file.filename == "":
            return jsonify(error="No file selected"), 400

        # Save the file
        filename = save_profile_picture(file, current_user.id)
        if not filename:
            return (
                jsonify(error="Invalid file type. Please use PNG, JPG, JPEG, or GIF."),
                400,
            )

        # Update user's profile picture
        current_user.profile_picture = filename
        db.session.commit()

        return (
            jsonify(
                message="Profile picture updated successfully!",
                profile_picture=f"/static/profile_pictures/{filename}",
            ),
            200,
        )
    except Exception as e:
        app.logger.error(f"Error uploading profile picture: {e}")
        db.session.rollback()
        return jsonify(error="An error occurred while uploading profile picture."), 500


@app.route("/api/request-password-reset", methods=["POST"])
def request_password_reset():
    """Request a password reset email."""
    try:
        data = request.get_json()
        if not data:
            return jsonify(error="No data provided"), 400

        email = data.get("email", "").lower().strip()
        user = User.query.filter_by(email=email).first()
        if user:
            send_password_reset_email(user.id, user.email, user.name)
        return (
            jsonify(
                message="If an account with that email exists, a reset link has been sent."
            ),
            200,
        )
    except Exception as e:
        app.logger.error(f"Error requesting password reset: {e}")
        return jsonify(error="An error occurred while processing your request."), 500


@app.route("/api/reset-password/<token>", methods=["POST"])
def reset_password(token):
    """Reset password using token."""
    try:
        reset_record = PasswordResetToken.query.filter_by(token=token).first()
        if not (reset_record and reset_record.expires_at > datetime.utcnow()):
            return jsonify(error="Invalid or expired token."), 400

        data = request.get_json()
        if not data:
            return jsonify(error="No data provided"), 400

        password = data.get("password")
        if not password:
            return jsonify(error="Password is required."), 400

        user = User.query.get(reset_record.user_id)
        if user:
            user.password_hash = generate_password_hash(password)
            db.session.delete(reset_record)
            db.session.commit()
            return jsonify(message="Password has been reset successfully."), 200
        else:
            return jsonify(error="User not found."), 404
    except Exception as e:
        app.logger.error(f"Error resetting password: {e}")
        db.session.rollback()
        return jsonify(error="An error occurred while resetting password."), 500


@app.route("/api/send-verification-email", methods=["POST"])
@login_required
def send_verification_email_api():
    """Send verification email to current user."""
    try:
        if current_user.is_verified:
            return jsonify(error="Your email is already verified."), 400

        success = send_verification_email(
            current_user.id, current_user.email, current_user.name
        )

        if success:
            return jsonify(message="Verification email sent successfully."), 200
        else:
            return (
                jsonify(
                    error="Failed to send verification email. Please check your email configuration."
                ),
                500,
            )
    except Exception as e:
        app.logger.error(f"Error sending verification email: {e}")
        return jsonify(error="An error occurred while sending verification email."), 500


@app.route("/api/check-auth")
def check_auth():
    """Check if user is authenticated and return user info."""
    if current_user.is_authenticated:
        app.logger.info(f"🔧 Check-auth for user: {current_user.email}")
        app.logger.info(f"🔧 User is_admin: {current_user.is_admin}")
        app.logger.info(f"🔧 User current_plan: {current_user.current_plan}")

        return jsonify(
            {
                "is_logged_in": True,
                "user": {
                    "name": current_user.name,
                    "email": current_user.email,
                    "is_verified": current_user.is_verified,
                    "current_plan": current_user.current_plan,
                    "is_admin": current_user.is_admin,
                    "profile_picture": (
                        f"/static/profile_pictures/{current_user.profile_picture}"
                        if current_user.profile_picture
                        else None
                    ),
                },
            }
        )
    return jsonify({"is_logged_in": False})


@app.route("/api/search", methods=["POST"])
def search():
    """Search for business leads."""
    try:
        data = request.get_json()
        if not data:
            return jsonify(error="No data provided"), 400

        business_type = data.get("business_type")
        radius_miles = data.get("radius", 3)  # Default to 3 miles
        max_reviews = data.get("max_reviews")

        try:
            radius_miles = int(radius_miles)
        except (ValueError, TypeError):
            radius_miles = 3  # Default to 3 if conversion fails

        # Convert miles to meters
        radius_meters = radius_miles * 1609.34

        try:
            if max_reviews:
                max_reviews = int(max_reviews)
            else:
                max_reviews = 100  # Default to 100 if not specified
        except (ValueError, TypeError):
            max_reviews = 100  # Default to 100 if conversion fails

        lat = data.get("lat")
        lng = data.get("lng")

        if not (lat and lng):
            if not (data.get("city") and data.get("state")):
                return (
                    jsonify({"error": "Either a map pin or city/state is required."}),
                    400,
                )

            city = data.get("city")
            state = data.get("state")
            location_query = f"{city}, {state}"

            # Log the geocoding attempt
            app.logger.info(f"Attempting to geocode: '{location_query}'")
            app.logger.info(f"API key available: {bool(app.config['GOOGLE_API_KEY'])}")
            app.logger.info(
                f"Environment variable GOOGLE_MAPS_API_KEY: "
                f"{bool(os.environ.get('GOOGLE_MAPS_API_KEY'))}"
            )

            coords_dict = get_coordinates(location_query, app.config["GOOGLE_API_KEY"])
            if not coords_dict:
                app.logger.error(f"Geocoding failed for: '{location_query}'")
                return (
                    jsonify(
                        error=f"Could not find coordinates for '{location_query}'. "
                        "Please check the city and state spelling."
                    ),
                    400,
                )
            lat, lng = coords_dict["lat"], coords_dict["lng"]
            app.logger.info(
                f"Successfully geocoded '{location_query}' to lat={lat}, lng={lng}"
            )

        # coords = (lat, lng)  # Removed unused variable

        if not business_type:
            return jsonify(error="Business type is required"), 400

        # --- GUEST SEARCH LIMITS ---
        # Track guest usage by IP address to prevent abuse
        if not current_user.is_authenticated or current_user.current_plan is None:
            guest_usage = get_or_create_guest_usage()

            if guest_usage:
                # Check if guest has exceeded daily limit (5 searches per day)
                if guest_usage.search_count >= 5:
                    return (
                        jsonify(
                            error="You've reached the daily limit of 5 free searches. "
                            "Please sign up or log in for unlimited searches."
                        ),
                        403,
                    )

                # Increment search count for this IP
                guest_usage.search_count += 1
                db.session.commit()
                max_results = 15
            else:
                # Fallback to session-based tracking if database tracking fails
                session.setdefault("guest_search_count", 0)
                if session["guest_search_count"] >= 5:
                    return (
                        jsonify(
                            error="Guest users are limited to 5 searches. "
                            "Please sign up or log in for more."
                        ),
                        403,
                    )
                session["guest_search_count"] += 1
                max_results = 15
        else:
            max_results = None  # No limit for authenticated users with a plan

        # Only update search count for authenticated users
        if current_user.is_authenticated:
            current_user.search_count = getattr(current_user, "search_count", 0) + 1
            db.session.commit()

        leads, center = search_places(
            lat,
            lng,
            business_type,
            radius_meters,
            app.config["GOOGLE_API_KEY"],
            max_reviews=max_reviews,
        )
        # Limit results for guests
        if max_results is not None:
            leads = leads[:max_results]
        session["last_search_results"] = leads

        # Track search action
        track_user_action("search", "search_page")

        return jsonify({"results": leads, "center": center})

    except Exception as e:
        app.logger.error(
            ("An error occurred during search: " + str(e)[:40] + str(e)[40:])
        )
        return jsonify(error="An unexpected error occurred during the search."), 500


@app.route("/download", methods=["POST"])
@login_required
def download():
    """Download search results as CSV or Excel file."""
    try:
        leads = session.get("last_search_results")
        if not leads:
            return "No leads to download.", 400

        file_format = request.form.get("format", "csv")
        if file_format == "csv":
            df = pd.DataFrame(leads)
            output = BytesIO()
            df.to_csv(output, index=False, encoding="utf-8")
            output.seek(0)
            # Track export action
            track_user_action("export", "csv_download")

            return send_file(
                output,
                mimetype="text/csv",
                as_attachment=True,
                download_name="leads.csv",
            )
        elif file_format == "xlsx":
            df = pd.DataFrame(leads)
            # Fix for pandas ExcelWriter with BytesIO - use a different approach
            workbook = Workbook()
            worksheet = workbook.active
            if worksheet:
                worksheet.title = "Leads"

                # Write headers
                for col, header in enumerate(df.columns, 1):
                    worksheet.cell(row=1, column=col, value=header)

                # Write data
                for row_idx, row_data in enumerate(df.values, 2):
                    for col_idx, value in enumerate(row_data, 1):
                        worksheet.cell(row=row_idx, column=col_idx, value=value)

            output = BytesIO()
            workbook.save(output)
            output.seek(0)
            # Track export action
            track_user_action("export", "xlsx_download")

            return send_file(
                output,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                as_attachment=True,
                download_name="leads.xlsx",
            )
        return "Invalid format", 400
    except Exception as e:
        app.logger.error(("Error downloading file: " + str(e)[:40] + str(e)[40:]))
        return "An error occurred while downloading the file.", 500


@app.route("/export-to-google-sheets", methods=["POST"])
@login_required
def export_to_google_sheets():
    """Export search results to Google Sheets."""
    try:
        leads = session.get("last_search_results")
        if not leads:
            return jsonify(error="No leads data to export."), 400

        gc = get_gspread_client()
        spreadsheet = gc.create(f"Leads - {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        worksheet = spreadsheet.get_worksheet(0)

        df = pd.DataFrame(leads)
        worksheet.update([df.columns.values.tolist()] + df.values.tolist())

        spreadsheet.share(current_user.email, perm_type="user", role="writer")

        # Track export action
        track_user_action("export", "google_sheets")

        return (
            jsonify(
                message="Successfully exported to Google Sheets!", url=spreadsheet.url
            ),
            200,
        )
    except Exception as e:
        app.logger.error(("Google Sheets export error: " + str(e)[:40] + str(e)[40:]))
        return jsonify(error=f"Failed to export to Google Sheets: {e}"), 500


@app.route("/api/settings", methods=["GET", "PUT"])
@login_required
def settings():
    """Get or update user settings."""
    try:
        user_settings = current_user.settings
        if not user_settings:
            user_settings = UserSettings()
            user_settings.user_id = current_user.id
            db.session.add(user_settings)
            db.session.commit()

        if request.method == "PUT":
            data = request.get_json()
            if not data:
                return jsonify(error="No data provided"), 400

            for key, value in data.items():
                if hasattr(user_settings, key):
                    setattr(user_settings, key, value)
            db.session.add(user_settings)
            db.session.commit()
            return jsonify(message="Settings updated successfully.")

        return jsonify(
            {
                "default_radius": user_settings.default_radius,
                "default_business_type": user_settings.default_business_type,
                "remember_last_search": user_settings.remember_last_search,
                "results_per_page": user_settings.results_per_page,
                "show_map_by_default": user_settings.show_map_by_default,
                "email_notifications": user_settings.email_notifications,
                "search_reminders": user_settings.search_reminders,
                "last_search_city": user_settings.last_search_city,
                "last_search_state": user_settings.last_search_state,
                "last_search_business_type": user_settings.last_search_business_type,
                "last_search_radius": user_settings.last_search_radius,
            }
        )
    except Exception as e:
        app.logger.error(("Error in settings: " + str(e)[:40] + str(e)[40:]))
        db.session.rollback()
        return jsonify(error="An error occurred while processing settings."), 500


@app.route("/api/change-password", methods=["POST"])
@login_required
def change_password():
    """Change user password."""
    try:
        data = request.get_json()
        if not data:
            return jsonify(error="No data provided"), 400

        if not current_user.check_password(data.get("current_password", "")):
            return jsonify(error="Invalid current password."), 400
        new_password = data.get("new_password")
        if not new_password or len(new_password) < 8:
            return (
                jsonify(error="New password must be at least 8 characters long."),
                400,
            )
        current_user.password_hash = generate_password_hash(new_password)
        db.session.commit()
        return jsonify(message="Password updated successfully.")
    except Exception as e:
        app.logger.error(f"Error changing password: {e}")
        db.session.rollback()
        return jsonify(error="An error occurred while changing password."), 500


@app.route("/api/export-data", methods=["GET"])
@login_required
def export_user_data():
    """Export user data as JSON."""
    try:
        user_data = {
            "profile": {
                "name": current_user.name,
                "email": current_user.email,
                "business": current_user.business,
                "phone": current_user.phone,
            },
            "settings": {
                "default_radius": (
                    current_user.settings.default_radius
                    if current_user.settings
                    else None
                ),
                "default_business_type": (
                    current_user.settings.default_business_type
                    if current_user.settings
                    else None
                ),
            },
        }

        output = BytesIO()
        output.write(json.dumps(user_data, indent=4).encode("utf-8"))
        output.seek(0)

        return send_file(
            output,
            mimetype="application/json",
            as_attachment=True,
            download_name="user_data.json",
        )
    except Exception as e:
        app.logger.error(f"Error exporting user data: {e}")
        return "An error occurred while exporting data.", 500


@app.route("/api/delete-account", methods=["POST"])
@login_required
def delete_account():
    """Delete user account."""
    try:
        data = request.get_json()
        if not data:
            return jsonify(error="No data provided"), 400

        if not current_user.check_password(data.get("password", "")):
            return jsonify(error="Invalid password."), 400

        # Optional: Cancel Stripe subscription before deleting
        if current_user.stripe_subscription_id:
            try:
                stripe.Subscription.delete(current_user.stripe_subscription_id)
            except Exception as e:
                app.logger.error(
                    f"Stripe subscription cancellation failed for user {current_user.id}: {e}"
                )
                # Decide if this should prevent account deletion

        db.session.delete(current_user)
        db.session.commit()
        logout_user()

        return jsonify(message="Your account has been permanently deleted.")
    except Exception as e:
        app.logger.error(f"Error deleting account: {e}")
        db.session.rollback()
        return jsonify(error="An error occurred while deleting your account."), 500


@app.route("/api/update-last-search", methods=["POST"])
@login_required
def update_last_search():
    """Update user's last search parameters."""
    try:
        data = request.get_json()
        if not data:
            return jsonify(error="No data provided"), 400

        user_settings = current_user.settings
        if not user_settings:
            user_settings = UserSettings()
            user_settings.user_id = current_user.id
            db.session.add(user_settings)

        user_settings.last_search_city = data.get("city")
        user_settings.last_search_state = data.get("state")
        user_settings.last_search_business_type = data.get("business_type")
        user_settings.last_search_radius = data.get("radius")
        db.session.add(user_settings)
        db.session.commit()
        return jsonify(success=True)
    except Exception as e:
        app.logger.error(f"Error updating last search: {e}")
        db.session.rollback()
        return jsonify(error="An error occurred while updating search parameters."), 500


@app.route("/api/create-checkout-session", methods=["POST"])
@login_required
def create_checkout_session():
    """Create Stripe checkout session."""
    try:
        data = request.get_json()
        if not data:
            return jsonify(error="No data provided"), 400

        # User must be authenticated to create checkout session
        customer_id = current_user.stripe_customer_id
        if not customer_id:
            # Create Stripe customer for authenticated user
            customer = stripe.Customer.create(
                email=current_user.email, name=current_user.name
            )
            current_user.stripe_customer_id = customer.id
            db.session.commit()
            customer_id = customer.id

        checkout_session = stripe.checkout.Session.create(
            customer=customer_id,
            payment_method_types=["card"],
            line_items=[{"price": data.get("priceId"), "quantity": 1}],
            mode="subscription",
            success_url=url_for("index", _external=True)
            + "?session_id={CHECKOUT_SESSION_ID}",
            cancel_url=url_for("pricing", _external=True),
        )
        return jsonify({"id": checkout_session.id})
    except Exception as e:
        app.logger.error(f"Error creating checkout session: {e}")
        return jsonify(error=str(e)), 403


@app.route("/api/create-portal-session", methods=["POST"])
@login_required
def create_portal_session():
    """Create Stripe customer portal session."""
    try:
        portal_session = stripe.billing_portal.Session.create(
            customer=current_user.stripe_customer_id,
            return_url=url_for("index", _external=True),
        )
        return jsonify({"url": portal_session.url})
    except Exception as e:
        app.logger.error(f"Error creating portal session: {e}")
        return jsonify(error=str(e)), 403


@app.route("/stripe-webhook", methods=["POST"])
def stripe_webhook():
    """Handle Stripe webhook events."""
    try:
        payload = request.get_data(as_text=True)
        sig_header = request.headers.get("Stripe-Signature")
        endpoint_secret = os.getenv("STRIPE_WEBHOOK_SECRET")

        try:
            event = stripe.Webhook.construct_event(payload, sig_header, endpoint_secret)
        except ValueError:
            return "Invalid payload", 400
        except Exception:
            return "Invalid signature", 400

        if (
            event["type"] == "customer.subscription.updated"
            or event["type"] == "customer.subscription.created"
        ):
            subscription = event["data"]["object"]
            customer_id = subscription["customer"]
            user = User.query.filter_by(stripe_customer_id=customer_id).first()
            if user:
                user.stripe_subscription_id = subscription["id"]
                user.current_plan = subscription["items"]["data"][0]["price"][
                    "lookup_key"
                ]
                db.session.commit()

        return "Success", 200
    except Exception as e:
        app.logger.error(f"Error processing Stripe webhook: {e}")
        return "Error processing webhook", 500


def get_gspread_client():
    """Get Google Sheets client."""
    scope = [
        "https://spreadsheets.google.com/feeds",
        "https://www.googleapis.com/auth/drive",
    ]
    creds_json = os.getenv("GOOGLE_SHEETS_CREDENTIALS_JSON")
    if not creds_json:
        raise ValueError("GOOGLE_SHEETS_CREDENTIALS_JSON environment variable not set")

    creds_dict = json.loads(creds_json)
    creds = Credentials.from_service_account_info(creds_dict, scopes=scope)
    return gspread.authorize(creds)


@app.route("/pricing")
def pricing():
    """Display pricing page."""
    # Get Stripe publishable key from environment
    stripe_publishable_key = os.getenv("STRIPE_PUBLISHABLE_KEY", "")

    # These would normally come from your Stripe dashboard
    # For now, using placeholder values
    basic_price_id = os.getenv("STRIPE_BASIC_PRICE_ID", "price_basic")
    premium_price_id = os.getenv("STRIPE_PREMIUM_PRICE_ID", "price_premium")
    platinum_price_id = os.getenv("STRIPE_PLATINUM_PRICE_ID", "price_platinum")

    # Track page visit
    track_page_visit("pricing")

    return render_template(
        "pricing.html",
        stripe_publishable_key=stripe_publishable_key,
        basic_price_id=basic_price_id,
        premium_price_id=premium_price_id,
        platinum_price_id=platinum_price_id,
    )


@app.route("/staff-registration")
def staff_registration():
    """Staff registration page"""
    return render_template("staff_registration.html")


def admin_required(f):
    """Decorator to require admin privileges."""

    @wraps(f)
    def decorated_function(*args, **kwargs):
        app.logger.info(f"🔧 Admin check for user: {current_user.email}")
        app.logger.info(f"🔧 User is_admin: {current_user.is_admin}")
        app.logger.info(f"🔧 User current_plan: {current_user.current_plan}")
        app.logger.info(f"🔧 User authenticated: {current_user.is_authenticated}")

        if not current_user.is_admin:
            app.logger.warning(f"❌ Admin access denied for user: {current_user.email}")
            flash("You do not have permission to access this page.", "danger")
            return redirect(url_for("index"))
        app.logger.info(f"✅ Admin access granted for user: {current_user.email}")
        return f(*args, **kwargs)

    return decorated_function


@app.route("/admin/dashboard")
@login_required
@admin_required
def admin_dashboard():
    """Admin dashboard page."""
    try:
        app.logger.info(f"Admin dashboard accessed by user: {current_user.email}")
        app.logger.info(f"User is_admin: {current_user.is_admin}")
        app.logger.info(f"User current_plan: {current_user.current_plan}")

        # Track admin dashboard visit
        track_page_visit("admin_dashboard")

        # Get analytics data for the last 30 days
        analytics_data = get_analytics_data(30)

        # Get user statistics
        total_users = User.query.count()
        active_users_today = (
            UserActivity.query.filter(
                db.func.date(UserActivity.created_at) == datetime.utcnow().date(),
                UserActivity.user_id.isnot(None),
            )
            .distinct(UserActivity.user_id)
            .count()
        )

        # Get plan distribution
        plan_distribution = (
            db.session.query(User.current_plan, db.func.count(User.id).label("count"))
            .group_by(User.current_plan)
            .all()
        )

        users = User.query.all()

        # Get AI teams data for the dashboard
        teams = AITeam.query.all()
        managers = AIManager.query.all()
        agents = AIAgent.query.all()

        return render_template(
            "admin_dashboard.html",
            users=users,
            analytics_data=analytics_data,
            total_users=total_users,
            active_users_today=active_users_today,
            plan_distribution=plan_distribution,
            teams=teams,
            managers=managers,
            agents=agents,
        )
    except Exception as e:
        app.logger.error(f"Error accessing admin dashboard: {e}")
        flash("An error occurred while loading the dashboard.", "danger")
        return redirect(url_for("index"))


@app.route("/restore-admin/<email>")
def restore_admin(email):
    """Temporary route to restore admin privileges for a user."""
    try:
        user = User.query.filter_by(email=email.lower().strip()).first()
        if user:
            user.is_admin = True
            user.current_plan = "admin"
            db.session.commit()
            return jsonify(message=f"Admin privileges restored for {email}"), 200
        else:
            return jsonify(error=f"User {email} not found"), 404
    except Exception as e:
        app.logger.error(f"Error restoring admin: {e}")
        return jsonify(error="An error occurred"), 500


@app.route("/temp-admin-login")
def temp_admin_login():
    """Temporary route to log in as admin for testing."""
    try:
        admin_user = User.query.filter_by(email="sdclark117@gmail.com").first()
        if admin_user:
            login_user(admin_user)
            app.logger.info(f"🔧 Temporary admin login for: {admin_user.email}")
            return redirect(url_for("index"))
        else:
            return "Admin user not found", 404
    except Exception as e:
        app.logger.error(f"❌ Temp admin login error: {e}")
        return "Error logging in as admin", 500


# Add a scheduled task to clean up expired tokens
_cleanup_counter = 0


@app.before_request
def before_request():
    """Run before each request to clean up expired tokens and reset guest usage occasionally."""
    global _cleanup_counter
    # Only clean up tokens occasionally (every 100 requests) to avoid performance impact
    _cleanup_counter += 1

    if _cleanup_counter % 100 == 0:
        cleanup_expired_tokens()
        reset_guest_usage_daily()
        update_daily_analytics()


# AI Teams Management API Routes
@app.route("/api/ai-teams/managers", methods=["GET", "POST"])
@login_required
@admin_required
def ai_managers():
    """Get all AI managers or create a new one."""
    if request.method == "GET":
        try:
            managers = AIManager.query.all()
            return (
                jsonify(
                    [
                        {
                            "id": manager.id,
                            "name": manager.name,
                            "email": manager.email,
                            "created_at": manager.created_at.isoformat(),
                            "team_id": manager.team.id if manager.team else None,
                            "team_name": manager.team.name if manager.team else None,
                        }
                        for manager in managers
                    ]
                ),
                200,
            )
        except Exception as e:
            app.logger.error(f"Error fetching AI managers: {e}")
            return jsonify(error="Failed to fetch managers"), 500

    elif request.method == "POST":
        try:
            data = request.get_json()
            if not data or not data.get("name") or not data.get("email"):
                return jsonify(error="Name and email are required"), 400

            # Check if email already exists
            if AIManager.query.filter_by(email=data["email"]).first():
                return jsonify(error="Manager with this email already exists"), 409

            manager = AIManager(name=data["name"], email=data["email"])
            db.session.add(manager)
            db.session.commit()

            app.logger.info(f"Created AI manager: {manager.name} ({manager.email})")
            return (
                jsonify(
                    {
                        "id": manager.id,
                        "name": manager.name,
                        "email": manager.email,
                        "created_at": manager.created_at.isoformat(),
                    }
                ),
                201,
            )
        except Exception as e:
            app.logger.error(f"Error creating AI manager: {e}")
            db.session.rollback()
            return jsonify(error="Failed to create manager"), 500


@app.route("/api/ai-teams/teams", methods=["GET", "POST"])
@login_required
@admin_required
def ai_teams():
    """Get all AI teams or create a new one."""
    if request.method == "GET":
        try:
            teams = AITeam.query.all()
            return (
                jsonify(
                    [
                        {
                            "id": team.id,
                            "name": team.name,
                            "created_at": team.created_at.isoformat(),
                            "manager_id": team.manager_id,
                            "manager_name": team.manager.name if team.manager else None,
                            "agent_count": len(team.agents),
                        }
                        for team in teams
                    ]
                ),
                200,
            )
        except Exception as e:
            app.logger.error(f"Error fetching AI teams: {e}")
            return jsonify(error="Failed to fetch teams"), 500

    elif request.method == "POST":
        try:
            data = request.get_json()
            if not data or not data.get("name"):
                return jsonify(error="Team name is required"), 400

            team = AITeam(name=data["name"])
            db.session.add(team)
            db.session.commit()

            app.logger.info(f"Created AI team: {team.name}")
            return (
                jsonify(
                    {
                        "id": team.id,
                        "name": team.name,
                        "created_at": team.created_at.isoformat(),
                    }
                ),
                201,
            )
        except Exception as e:
            app.logger.error(f"Error creating AI team: {e}")
            db.session.rollback()
            return jsonify(error="Failed to create team"), 500


@app.route("/api/ai-teams/agents", methods=["GET", "POST"])
@login_required
@admin_required
def ai_agents():
    """Get all AI agents or create a new one."""
    if request.method == "GET":
        try:
            agents = AIAgent.query.all()
            return (
                jsonify(
                    [
                        {
                            "id": agent.id,
                            "name": agent.name,
                            "specialty": agent.specialty,
                            "status": agent.status,
                            "created_at": agent.created_at.isoformat(),
                            "team_id": agent.team_id,
                            "team_name": agent.team.name if agent.team else None,
                        }
                        for agent in agents
                    ]
                ),
                200,
            )
        except Exception as e:
            app.logger.error(f"Error fetching AI agents: {e}")
            return jsonify(error="Failed to fetch agents"), 500

    elif request.method == "POST":
        try:
            data = request.get_json()
            if not data or not data.get("name") or not data.get("specialty"):
                return jsonify(error="Name and specialty are required"), 400

            # Handle custom specialty
            specialty = data["specialty"]
            if specialty == "Custom" and data.get("custom_specialty"):
                specialty = data["custom_specialty"]

            agent = AIAgent(
                name=data["name"], specialty=specialty, team_id=data.get("team_id")
            )
            db.session.add(agent)
            db.session.commit()

            app.logger.info(
                f"Created AI agent: {agent.name} (Specialty: {agent.specialty})"
            )
            return (
                jsonify(
                    {
                        "id": agent.id,
                        "name": agent.name,
                        "specialty": agent.specialty,
                        "status": agent.status,
                        "created_at": agent.created_at.isoformat(),
                        "team_id": agent.team_id,
                    }
                ),
                201,
            )
        except Exception as e:
            app.logger.error(f"Error creating AI agent: {e}")
            db.session.rollback()
            return jsonify(error="Failed to create agent"), 500


@app.route("/api/ai-teams/assign-manager/<int:team_id>", methods=["POST"])
@login_required
@admin_required
def assign_manager_to_team(team_id):
    """Assign a manager to a team."""
    try:
        data = request.get_json()
        if not data or not data.get("manager_id"):
            return jsonify(error="Manager ID is required"), 400

        team = AITeam.query.get(team_id)
        if not team:
            return jsonify(error="Team not found"), 404

        manager = AIManager.query.get(data["manager_id"])
        if not manager:
            return jsonify(error="Manager not found"), 404

        team.manager_id = manager.id
        db.session.commit()

        app.logger.info(f"Assigned manager {manager.name} to team {team.name}")
        return jsonify(message="Manager assigned successfully"), 200
    except Exception as e:
        app.logger.error(f"Error assigning manager to team: {e}")
        db.session.rollback()
        return jsonify(error="Failed to assign manager"), 500


@app.route("/api/ai-teams/specialties", methods=["GET"])
@login_required
@admin_required
def get_ai_specialties():
    """Get list of available AI specialties."""
    specialties = [
        "Marketing",
        "Cold Calling",
        "Social Media Management",
        "Email Campaigns",
        "Lead Generation",
        "Content Creation",
        "SEO",
        "Analytics",
        "Customer Support",
        "Sales",
    ]
    return jsonify(specialties), 200


@app.route("/api/admin/users", methods=["GET"])
@login_required
@admin_required
def get_all_users():
    """Get all users for staff management"""
    try:
        users = User.query.all()
        user_list = []

        for user in users:
            user_data = {
                "id": user.id,
                "name": user.name,
                "email": user.email,
                "is_admin": user.is_admin,
                "is_support": user.is_support,
                "is_technical": user.is_technical,
                "staff_role": user.staff_role,
                "current_plan": user.current_plan,
                "is_verified": user.is_verified,
                "created_at": user.created_at.isoformat() if user.created_at else None,
            }
            user_list.append(user_data)

        return jsonify(user_list)
    except Exception as e:
        app.logger.error(f"Error getting users: {e}")
        return jsonify({"error": "Failed to get users"}), 500


@app.route("/api/admin/staff", methods=["POST"])
@login_required
@admin_required
def create_staff_member():
    """Create a new staff member"""
    try:
        data = request.get_json()
        name = data.get("name")
        email = data.get("email")
        role = data.get("role")  # 'support' or 'technical'

        if not all([name, email, role]):
            return jsonify({"error": "Name, email, and role are required"}), 400

        if role not in ["support", "technical"]:
            return jsonify({"error": "Role must be 'support' or 'technical'"}), 400

        # Check if user already exists
        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            return jsonify({"error": "User with this email already exists"}), 400

        # Create new staff member
        user = User(
            name=name,
            email=email,
            password_hash=generate_password_hash(
                "temp_password_123"
            ),  # Temporary password
            is_support=(role == "support"),
            is_technical=(role == "technical"),
            staff_role=role,
            is_verified=True,  # Staff members are auto-verified
        )

        db.session.add(user)
        db.session.commit()

        return (
            jsonify(
                {"message": "Staff member created successfully", "user_id": user.id}
            ),
            201,
        )

    except Exception as e:
        app.logger.error(f"Error creating staff member: {e}")
        return jsonify({"error": "Failed to create staff member"}), 500


@app.route("/api/admin/staff/<int:user_id>", methods=["PUT"])
@login_required
@admin_required
def update_staff_member(user_id):
    """Update staff member role"""
    try:
        data = request.get_json()
        new_role = data.get("role")

        if not new_role or new_role not in ["support", "technical"]:
            return jsonify({"error": "Valid role (support/technical) is required"}), 400

        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Update staff role
        user.is_support = new_role == "support"
        user.is_technical = new_role == "technical"
        user.staff_role = new_role

        db.session.commit()

        return jsonify(
            {
                "message": "Staff member updated successfully",
                "user_id": user.id,
                "new_role": new_role,
            }
        )

    except Exception as e:
        app.logger.error(f"Error updating staff member: {e}")
        return jsonify({"error": "Failed to update staff member"}), 500


@app.route("/api/admin/staff/<int:user_id>", methods=["DELETE"])
@login_required
@admin_required
def remove_staff_member(user_id):
    """Remove staff member status (convert to regular user)"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Remove staff status
        user.is_support = False
        user.is_technical = False
        user.staff_role = None

        db.session.commit()

        return jsonify(
            {"message": "Staff member removed successfully", "user_id": user.id}
        )

    except Exception as e:
        app.logger.error(f"Error removing staff member: {e}")
        return jsonify({"error": "Failed to remove staff member"}), 500


@app.route("/api/admin/users/<int:user_id>", methods=["DELETE"])
@login_required
@admin_required
def delete_user(user_id):
    """Delete a user account"""
    try:
        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Prevent deleting admin accounts
        if user.is_admin:
            return jsonify({"error": "Cannot delete admin accounts"}), 400

        db.session.delete(user)
        db.session.commit()

        return jsonify({"message": "User deleted successfully"}), 200
    except Exception as e:
        app.logger.error(f"Error deleting user: {e}")
        db.session.rollback()
        return jsonify({"error": "Failed to delete user"}), 500


@app.route("/api/admin/promote-user/<int:user_id>", methods=["POST"])
@login_required
@admin_required
def promote_user_to_staff(user_id):
    """Promote a regular user to staff"""
    try:
        data = request.get_json()
        role = data.get("role")

        if not role or role not in ["support", "technical"]:
            return jsonify({"error": "Valid role (support/technical) is required"}), 400

        user = User.query.get(user_id)
        if not user:
            return jsonify({"error": "User not found"}), 404

        # Promote to staff
        user.is_support = role == "support"
        user.is_technical = role == "technical"
        user.staff_role = role

        db.session.commit()

        return jsonify(
            {
                "message": "User promoted to staff successfully",
                "user_id": user.id,
                "new_role": role,
            }
        )

    except Exception as e:
        app.logger.error(f"Error promoting user: {e}")
        return jsonify({"error": "Failed to promote user"}), 500


@app.route("/api/admin/generate-access-code", methods=["POST"])
@login_required
@admin_required
def generate_access_code():
    """Generate an access code for staff registration."""
    try:
        data = request.get_json()
        if not data:
            return jsonify(error="No data provided"), 400

        staff_role = data.get("staff_role", "").strip()
        is_support = data.get("is_support", False)
        is_technical = data.get("is_technical", False)

        if not staff_role:
            return jsonify(error="Staff role is required"), 400

        # Generate a unique access code
        access_code = generate_token()[:8].upper()  # 8-character uppercase code

        # Create access code record
        access_code_record = StaffAccessCode(
            code=access_code,
            staff_role=staff_role,
            is_support=is_support,
            is_technical=is_technical,
            created_by=current_user.id,
            expires_at=datetime.utcnow() + timedelta(hours=24),
        )
        db.session.add(access_code_record)
        db.session.commit()

        return (
            jsonify(
                message="Access code generated successfully",
                access_code=access_code,
                expires_at=access_code_record.expires_at.isoformat(),
            ),
            201,
        )
    except Exception as e:
        app.logger.error(f"Error generating access code: {e}")
        db.session.rollback()
        return jsonify(error="An error occurred while generating access code"), 500


@app.route("/api/register-with-access-code", methods=["POST"])
def register_with_access_code():
    """Register a new staff member using an access code."""
    try:
        data = request.get_json()
        if not data:
            return jsonify(error="No data provided"), 400

        email = data.get("email", "").lower().strip()
        name = data.get("name", "").strip()
        password = data.get("password", "").strip()
        access_code = data.get("access_code", "").strip().upper()

        if not all([email, name, password, access_code]):
            return jsonify(error="All fields are required"), 400

        # Validate access code
        access_code_record = (
            StaffAccessCode.query.filter_by(code=access_code, is_used=False)
            .filter(StaffAccessCode.expires_at > datetime.utcnow())
            .first()
        )

        if not access_code_record:
            return jsonify(error="Invalid or expired access code"), 400

        # Check if user already exists
        existing_user = User.query.filter_by(email=email).first()
        if existing_user:
            return jsonify(error="User with this email already exists"), 400

        # Create new staff member
        staff_member = User(
            email=email,
            name=name,
            password_hash=generate_password_hash(password),
            is_support=access_code_record.is_support,
            is_technical=access_code_record.is_technical,
            staff_role=access_code_record.staff_role,
            is_verified=True,  # Staff members are pre-verified
        )
        db.session.add(staff_member)

        # Mark access code as used
        access_code_record.is_used = True
        access_code_record.used_by = staff_member.id
        access_code_record.used_at = datetime.utcnow()

        db.session.commit()

        return (
            jsonify(
                message="Staff member registered successfully",
                staff_member={
                    "id": staff_member.id,
                    "email": staff_member.email,
                    "name": staff_member.name,
                    "staff_role": staff_member.staff_role,
                    "is_support": staff_member.is_support,
                    "is_technical": staff_member.is_technical,
                },
            ),
            201,
        )
    except Exception as e:
        app.logger.error(f"Error registering with access code: {e}")
        db.session.rollback()
        return jsonify(error="An error occurred while registering"), 500


# AI Agent Functionality Systems
class AILeadGenerationSystem:
    """AI system for lead generation tasks"""

    def __init__(self):
        self.linkedin_api_key = os.environ.get("LINKEDIN_API_KEY")
        self.crunchbase_api_key = os.environ.get("CRUNCHBASE_API_KEY")
        self.leads_database = {}

    def find_prospects(self, industry, company_size, location):
        """Find new prospects based on criteria"""
        # Simulate LinkedIn and Crunchbase API calls
        prospects = [
            {
                "name": "TechCorp Solutions",
                "industry": "Technology",
                "size": "50-200 employees",
                "location": "San Francisco, CA",
                "contact": "john.doe@techcorp.com",
                "linkedin": "linkedin.com/in/johndoe",
                "score": 85,
            },
            {
                "name": "Innovate Labs",
                "industry": "Software",
                "size": "10-50 employees",
                "location": "Austin, TX",
                "contact": "sarah.smith@innovatelabs.com",
                "linkedin": "linkedin.com/in/sarahsmith",
                "score": 92,
            },
        ]
        return prospects

    def qualify_lead(self, prospect_data):
        """Qualify a lead based on BANT criteria"""
        score = 0
        if prospect_data.get("budget"):
            score += 25
        if prospect_data.get("authority"):
            score += 25
        if prospect_data.get("need"):
            score += 25
        if prospect_data.get("timeline"):
            score += 25
        return score >= 75

    def generate_outreach_sequence(self, prospect):
        """Generate personalized outreach sequence"""
        return {
            "email_1": f"Hi {prospect['name']}, I noticed your company's work in {prospect['industry']}...",
            "email_2": f"Following up on my previous email about {prospect['company']}...",
            "email_3": f"Final attempt to connect regarding {prospect['company']}...",
            "linkedin_message": f"Hi {prospect['name']}, I'd love to connect and discuss...",
            "call_script": f"Hi {prospect['name']}, this is [Name] calling from [Company]...",
        }


class AIColdCallingSystem:
    """AI system for cold calling tasks"""

    def __init__(self):
        self.call_scripts = {}
        self.call_metrics = {
            "total_calls": 0,
            "successful_calls": 0,
            "appointments_set": 0,
            "conversion_rate": 0.0,
        }

    def generate_call_script(self, prospect_data, product_info):
        """Generate personalized call script"""
        return {
            "opening": "Hi {}, this is [Name] calling from [Company]...".format(
                prospect_data["name"]
            ),
            "pain_point": "I understand companies like yours in {} are facing...".format(
                prospect_data["industry"]
            ),
            "solution": "Our solution has helped similar companies achieve...",
            "question": "Would you be interested in a 15-minute call to discuss how we could help?",
            "objection_handlers": {
                "not_interested": "I understand. Many people feel that way initially...",
                "no_time": "I respect your time. This would only take 15 minutes...",
                "send_info": "I'd be happy to send some information, but I'd also like to...",
            },
        }

    def track_call_result(self, call_id, result, notes):
        """Track call results and update metrics"""
        self.call_metrics["total_calls"] += 1
        if result == "success":
            self.call_metrics["successful_calls"] += 1
        if result == "appointment":
            self.call_metrics["appointments_set"] += 1

        self.call_metrics["conversion_rate"] = (
            self.call_metrics["successful_calls"] / self.call_metrics["total_calls"]
        )
        return self.call_metrics


class AISalesOutreachSystem:
    """AI system for sales outreach and follow-up"""

    def __init__(self):
        self.email_templates = {}
        self.follow_up_sequences = {}
        self.proposal_templates = {}

    def create_proposal(self, prospect_data, solution_details):
        """Create personalized sales proposal"""
        return {
            "executive_summary": "Based on our analysis of {}...".format(
                prospect_data["company"]
            ),
            "problem_statement": "Companies in {} face challenges with...".format(
                prospect_data["industry"]
            ),
            "solution_overview": "Our {} solution provides...".format(
                solution_details["product"]
            ),
            "value_proposition": "This will deliver {} ROI within 6 months...".format(
                solution_details["roi"]
            ),
            "implementation_plan": "Phase 1: {}...".format(
                solution_details["timeline"]
            ),
            "investment": "Total investment: ${}...".format(solution_details["price"]),
            "next_steps": "Schedule a technical review and contract signing...",
        }

    def generate_follow_up_sequence(self, prospect, last_interaction):
        """Generate automated follow-up sequence"""
        sequence = []
        days_since = (datetime.utcnow() - last_interaction).days

        if days_since == 1:
            sequence.append(
                {
                    "type": "email",
                    "subject": "Quick follow-up from our call",
                    "content": "Hi {}, thanks for taking my call yesterday...".format(
                        prospect["name"]
                    ),
                }
            )
        elif days_since == 3:
            sequence.append(
                {
                    "type": "email",
                    "subject": "Thought you might find this interesting",
                    "content": "Hi {}, I came across this article about {}...".format(
                        prospect["name"], prospect["industry"]
                    ),
                }
            )
        elif days_since == 7:
            sequence.append(
                {
                    "type": "call",
                    "script": "Hi {}, I wanted to follow up on our previous discussion...".format(
                        prospect["name"]
                    ),
                }
            )

        return sequence


class AISocialMediaSystem:
    """AI system for social media management"""

    def __init__(self):
        self.platforms = ["linkedin", "twitter", "instagram", "facebook"]
        self.content_calendar = {}
        self.engagement_metrics = {}

    def create_content_calendar(self, themes, frequency):
        """Create content calendar for the month"""
        calendar = {}
        content_types = [
            "educational",
            "promotional",
            "engagement",
            "thought_leadership",
        ]

        for week in range(1, 5):
            calendar[f"week_{week}"] = []
            for day in range(1, 8):
                content_type = content_types[day % len(content_types)]
                calendar[f"week_{week}"].append(
                    {
                        "day": day,
                        "type": content_type,
                        "theme": themes[content_type],
                        "platform": self.platforms[day % len(self.platforms)],
                        "content": self.generate_content(
                            content_type, themes[content_type]
                        ),
                    }
                )

        return calendar

    def generate_content(self, content_type, theme):
        """Generate social media content"""
        templates = {
            "educational": f"Did you know? {theme} can significantly impact your business. Here's how...",
            "promotional": f"🚀 Excited to share our latest {theme} solution! Learn more...",
            "engagement": f"What's your biggest challenge with {theme}? Share your thoughts below! 👇",
            "thought_leadership": f"The future of {theme} is evolving rapidly. Here's what you need to know...",
        }
        return templates.get(content_type, "Great insights on {}!".format(theme))

    def analyze_engagement(self, post_data):
        """Analyze post engagement and provide insights"""
        engagement_rate = (
            post_data["likes"] + post_data["comments"] + post_data["shares"]
        ) / post_data["reach"]
        return {
            "engagement_rate": engagement_rate,
            "best_performing_time": "9:00 AM - 11:00 AM",
            "optimal_content_type": "educational",
            "recommendations": [
                "Post more educational content",
                "Increase video content by 25%",
                "Engage with followers within 1 hour",
            ],
        }


class AIEmailCampaignSystem:
    """AI system for email campaign management"""

    def __init__(self):
        self.campaign_templates = {}
        self.segment_data = {}
        self.ab_test_results = {}

    def create_email_campaign(self, campaign_type, target_audience):
        """Create email campaign with A/B testing"""
        templates = {
            "welcome_series": {
                "subject_a": "Welcome to [Company] - Let's get started!",
                "subject_b": "Welcome! Here's your exclusive guide",
                "content_a": "Hi [Name], welcome to [Company]! We're excited to help you...",
                "content_b": "Welcome to [Company], [Name]! We've prepared a special guide just for you...",
            },
            "nurture_sequence": {
                "subject_a": "5 ways to improve your [Industry] results",
                "subject_b": "Quick tip: Boost your [Industry] performance",
                "content_a": "Hi [Name], here are 5 proven strategies to improve your [Industry] results...",
                "content_b": "Hi [Name], here's a quick tip that can immediately boost your [Industry] performance...",
            },
            "promotional": {
                "subject_a": "Limited time: 25% off [Product]",
                "subject_b": "Exclusive offer just for you, [Name]",
                "content_a": "Hi [Name], we're offering 25% off [Product] for a limited time...",
                "content_b": "Hi [Name], as a valued customer, we're offering you an exclusive discount...",
            },
        }

        return templates.get(campaign_type, templates["nurture_sequence"])

    def segment_audience(self, subscriber_data):
        """Segment email audience based on behavior and demographics"""
        segments = {
            "high_engagement": [],
            "medium_engagement": [],
            "low_engagement": [],
            "new_subscribers": [],
            "inactive": [],
        }

        for subscriber in subscriber_data:
            if subscriber["days_since_signup"] <= 7:
                segments["new_subscribers"].append(subscriber)
            elif subscriber["open_rate"] >= 0.3:
                segments["high_engagement"].append(subscriber)
            elif subscriber["open_rate"] >= 0.15:
                segments["medium_engagement"].append(subscriber)
            elif subscriber["days_since_open"] > 30:
                segments["inactive"].append(subscriber)
            else:
                segments["low_engagement"].append(subscriber)

        return segments

    def analyze_campaign_performance(self, campaign_data):
        """Analyze email campaign performance"""
        return {
            "open_rate": campaign_data["opens"] / campaign_data["sent"],
            "click_rate": campaign_data["clicks"] / campaign_data["sent"],
            "conversion_rate": campaign_data["conversions"] / campaign_data["sent"],
            "unsubscribe_rate": campaign_data["unsubscribes"] / campaign_data["sent"],
            "revenue_per_email": campaign_data["revenue"] / campaign_data["sent"],
            "recommendations": [
                "Test different subject lines",
                "Improve email timing",
                "Segment audience more granularly",
            ],
        }


class AIContentCreationSystem:
    """AI system for content creation and management"""

    def __init__(self):
        self.content_templates = {}
        self.seo_keywords = {}
        self.content_calendar = {}

    def generate_blog_post(self, topic, target_keywords, word_count=800):
        """Generate SEO-optimized blog post"""
        outline = {
            "title": "Complete Guide to {} in 2024".format(topic),
            "meta_description": (
                "Learn everything about {} and how it can benefit your business. "
                "Expert insights and actionable tips."
            ).format(topic),
            "introduction": "{} has become increasingly important for businesses looking to...".format(
                topic
            ),
            "sections": [
                "What is {}?".format(topic),
                "Why {} matters for your business".format(topic),
                "Best practices for implementing {}".format(topic),
                "Common mistakes to avoid",
                "Success stories and case studies",
                "Next steps and conclusion",
            ],
            "target_keywords": target_keywords,
            "estimated_read_time": "{} minutes".format(word_count // 200),
        }

        return outline

    def create_social_media_graphics(self, content_type, brand_guidelines):
        """Generate social media graphics specifications"""
        graphics = {
            "linkedin_post": {
                "dimensions": "1200x628px",
                "style": "Professional, clean, brand colors",
                "elements": ["Logo", "Headline", "CTA button", "Brand colors"],
            },
            "instagram_post": {
                "dimensions": "1080x1080px",
                "style": "Modern, engaging, visual hierarchy",
                "elements": [
                    "Eye-catching image",
                    "Overlay text",
                    "Hashtags",
                    "Brand watermark",
                ],
            },
            "twitter_post": {
                "dimensions": "1200x675px",
                "style": "Simple, readable, mobile-friendly",
                "elements": ["Clear text", "Minimal design", "Brand colors"],
            },
        }

        return graphics.get(content_type, graphics["linkedin_post"])

    def optimize_content_for_seo(self, content, target_keywords):
        """Optimize content for search engines"""
        optimization = {
            "keyword_density": {},
            "heading_structure": ["H1", "H2", "H3"],
            "meta_tags": {
                "title": "{} - Complete Guide | [Company]".format(target_keywords[0]),
                "description": (
                    "Learn about {} and how it can help your business. "
                    "Expert insights and actionable tips."
                ).format(target_keywords[0]),
                "keywords": ", ".join(target_keywords),
            },
            "internal_links": [
                "Link to related {} content".format(target_keywords[1]),
                "Link to {} case study".format(target_keywords[2]),
            ],
            "recommendations": [
                "Include more long-tail keywords",
                "Add more internal links",
                "Improve heading structure",
            ],
        }

        return optimization


class AISEOSystem:
    """AI system for SEO and analytics"""

    def __init__(self):
        self.keyword_tracker = {}
        self.rankings_database = {}
        self.analytics_data = {}

    def analyze_website_performance(self, domain):
        """Analyze website SEO performance"""
        return {
            "page_speed": {
                "mobile": 85,
                "desktop": 92,
                "recommendations": [
                    "Optimize images",
                    "Minimize CSS/JS",
                    "Enable compression",
                ],
            },
            "technical_seo": {
                "indexed_pages": 156,
                "broken_links": 3,
                "missing_meta": 12,
                "recommendations": [
                    "Fix broken links",
                    "Add meta descriptions",
                    "Improve site structure",
                ],
            },
            "keyword_rankings": {
                "top_10": 23,
                "top_50": 45,
                "top_100": 67,
                "opportunities": [
                    "Long-tail keywords",
                    "Local SEO",
                    "Featured snippets",
                ],
            },
            "traffic_analysis": {
                "organic_traffic": 23400,
                "organic_conversions": 156,
                "conversion_rate": 0.67,
                "trend": "+12% month over month",
            },
        }

    def generate_keyword_research(self, industry, target_location):
        """Generate keyword research and opportunities"""
        keywords = {
            "primary_keywords": [
                {
                    "keyword": "{} solutions".format(industry),
                    "volume": 1200,
                    "difficulty": 45,
                },
                {
                    "keyword": "best {} company".format(industry),
                    "volume": 890,
                    "difficulty": 52,
                },
                {
                    "keyword": "{} services near me".format(industry),
                    "volume": 650,
                    "difficulty": 38,
                },
            ],
            "long_tail_keywords": [
                {
                    "keyword": "how to choose {} provider".format(industry),
                    "volume": 320,
                    "difficulty": 28,
                },
                {
                    "keyword": "{} cost comparison".format(industry),
                    "volume": 210,
                    "difficulty": 35,
                },
                {
                    "keyword": "{} benefits for small business".format(industry),
                    "volume": 180,
                    "difficulty": 25,
                },
            ],
            "local_keywords": [
                {
                    "keyword": "{} {}".format(industry, target_location),
                    "volume": 450,
                    "difficulty": 42,
                },
                {
                    "keyword": "{} near {}".format(industry, target_location),
                    "volume": 380,
                    "difficulty": 38,
                },
            ],
        }

        return keywords

    def create_seo_optimization_plan(self, current_performance):
        """Create SEO optimization action plan"""
        return {
            "immediate_actions": [
                "Fix 3 broken links identified",
                "Add meta descriptions to 12 pages",
                "Optimize 5 high-traffic pages",
            ],
            "short_term_goals": [
                "Improve page speed by 15%",
                "Target 10 new long-tail keywords",
                "Create 5 pillar content pieces",
            ],
            "long_term_strategy": [
                "Build 50 quality backlinks",
                "Optimize for voice search",
                "Implement schema markup",
            ],
            "expected_results": {
                "traffic_increase": "+25% in 3 months",
                "ranking_improvements": "+15 positions average",
                "conversion_boost": "+18% organic conversions",
            },
        }


class AIPPCSystem:
    """AI system for PPC and advertising management"""

    def __init__(self):
        self.campaign_data = {}
        self.ad_performance = {}
        self.budget_allocations = {}

    def create_ppc_campaign(self, campaign_type, target_audience, budget):
        """Create PPC campaign with targeting and optimization"""
        campaign = {
            "campaign_name": "{} Campaign - {}".format(
                campaign_type.title(), target_audience
            ),
            "budget": budget,
            "platforms": ["Google Ads", "Facebook Ads", "LinkedIn Ads"],
            "targeting": {
                "demographics": {
                    "age": "25-54",
                    "gender": "All",
                    "income": "$50K+",
                    "education": "Bachelor's degree+",
                },
                "interests": [
                    "Business software",
                    "Marketing",
                    "Technology",
                    "Entrepreneurship",
                ],
                "behaviors": [
                    "Frequent online shoppers",
                    "Business decision makers",
                    "Technology early adopters",
                ],
            },
            "ad_groups": [
                {
                    "name": "Primary Keywords",
                    "keywords": ["business software", "marketing tools", "automation"],
                    "bids": [2.50, 3.20, 2.80],
                },
                {
                    "name": "Long-tail Keywords",
                    "keywords": [
                        "best marketing software for small business",
                        "automation tools for entrepreneurs",
                    ],
                    "bids": [1.80, 2.10],
                },
            ],
            "ad_copy": {
                "headlines": [
                    "Transform Your Business Today",
                    "Automate Your Marketing",
                    "Save 10 Hours Per Week",
                ],
                "descriptions": [
                    "Streamline your workflow with our powerful automation tools. Start your free trial today!",
                    "Join 10,000+ businesses using our platform. Get started in minutes.",
                ],
            },
        }

        return campaign

    def optimize_campaign_performance(self, campaign_data):
        """Optimize PPC campaign based on performance data"""
        optimizations = {
            "bid_adjustments": {
                "high_performing_keywords": "+20%",
                "low_performing_keywords": "-15%",
                "mobile_users": "+10%",
                "weekend_traffic": "+25%",
            },
            "ad_copy_tests": [
                "Test new headlines focusing on benefits",
                "A/B test different call-to-actions",
                "Experiment with different ad formats",
            ],
            "budget_reallocation": {
                "high_roi_campaigns": "+30% budget",
                "low_roi_campaigns": "-20% budget",
                "new_testing_budget": "10% of total",
            },
            "targeting_refinements": [
                "Exclude low-converting audiences",
                "Add similar audiences to high-performers",
                "Test new geographic markets",
            ],
        }

        return optimizations

    def analyze_campaign_roi(self, campaign_metrics):
        """Analyze campaign ROI and provide insights"""
        roi_analysis = {
            "overall_roi": campaign_metrics["revenue"] / campaign_metrics["cost"],
            "conversion_rate": campaign_metrics["conversions"]
            / campaign_metrics["clicks"],
            "cost_per_acquisition": campaign_metrics["cost"]
            / campaign_metrics["conversions"],
            "lifetime_value": campaign_metrics["lifetime_value"]
            / campaign_metrics["conversions"],
            "recommendations": [
                "Increase bids on high-converting keywords",
                "Pause low-performing ad groups",
                "Test new ad copy variations",
                "Expand to similar audiences",
            ],
            "forecast": {
                "projected_revenue": campaign_metrics["revenue"] * 1.25,
                "projected_roi": (campaign_metrics["revenue"] * 1.25)
                / campaign_metrics["cost"],
                "budget_recommendation": campaign_metrics["cost"] * 1.15,
            },
        }

        return roi_analysis


class AIBrandStrategySystem:
    """AI system for brand strategy and management"""

    def __init__(self):
        self.brand_guidelines = {}
        self.brand_performance = {}
        self.competitive_analysis = {}

    def develop_brand_strategy(self, company_data, target_audience):
        """Develop comprehensive brand strategy"""
        strategy = {
            "brand_positioning": {
                "value_proposition": "The leading {} solution for {}".format(
                    company_data["industry"], target_audience["primary"]
                ),
                "differentiators": [
                    "Advanced automation capabilities",
                    "Superior customer support",
                    "Proven ROI guarantee",
                ],
                "personality": "Professional, innovative, trustworthy, approachable",
            },
            "brand_identity": {
                "logo_guidelines": {
                    "primary_logo": "Full color on white background",
                    "secondary_logo": "White version for dark backgrounds",
                    "minimum_size": "1 inch width",
                    "clear_space": "Equal to logo height",
                },
                "color_palette": {
                    "primary": "#2563EB",
                    "secondary": "#10B981",
                    "accent": "#F59E0B",
                    "neutral": "#6B7280",
                },
                "typography": {
                    "headings": "Inter Bold",
                    "body_text": "Inter Regular",
                    "accent_text": "Inter Medium",
                },
            },
            "messaging_framework": {
                "core_message": "Empowering {} to achieve more with intelligent automation".format(
                    target_audience["primary"]
                ),
                "key_messages": [
                    "Simplify complex workflows",
                    "Drive measurable results",
                    "Scale with confidence",
                ],
                "tone_of_voice": "Professional yet approachable, confident but not arrogant",
            },
            "brand_guidelines": {
                "do": [
                    "Use brand colors consistently",
                    "Maintain professional tone",
                    "Focus on customer benefits",
                    "Show real results and testimonials",
                ],
                "dont": [
                    "Use competitor names",
                    "Make unrealistic claims",
                    "Use generic stock photos",
                    "Ignore brand voice guidelines",
                ],
            },
        }

        return strategy

    def conduct_brand_audit(self, current_brand_data):
        """Conduct comprehensive brand audit"""
        audit = {
            "brand_awareness": {
                "unaided_recall": 23,
                "aided_recall": 67,
                "brand_recognition": 78,
                "industry_rank": 4,
            },
            "brand_perception": {
                "trust_score": 8.2,
                "quality_score": 8.5,
                "innovation_score": 8.8,
                "customer_service_score": 9.1,
            },
            "competitive_positioning": {
                "market_share": 12,
                "competitive_advantage": "Superior customer support",
                "differentiation_score": 8.3,
            },
            "recommendations": [
                "Increase brand awareness through thought leadership",
                "Strengthen differentiation messaging",
                "Improve customer testimonials visibility",
                "Develop influencer partnerships",
            ],
        }

        return audit

    def create_brand_campaign(self, campaign_objective, target_audience):
        """Create brand awareness campaign"""
        campaign = {
            "campaign_name": "Brand Awareness - {}".format(campaign_objective),
            "objectives": [
                "Increase brand awareness by 25%",
                "Improve brand perception scores",
                "Generate 500 new brand mentions",
            ],
            "channels": {
                "social_media": {
                    "platforms": ["LinkedIn", "Twitter", "Instagram"],
                    "content_types": [
                        "Thought leadership",
                        "Behind-the-scenes",
                        "Customer stories",
                    ],
                    "frequency": "3-5 posts per week",
                },
                "content_marketing": {
                    "blog_posts": "2 per week",
                    "whitepapers": "1 per month",
                    "webinars": "1 per quarter",
                },
                "public_relations": {
                    "press_releases": "1 per month",
                    "media_outreach": "10 targeted journalists",
                    "industry_events": "3 speaking opportunities",
                },
            },
            "success_metrics": {
                "brand_awareness": "25% increase",
                "social_mentions": "+40%",
                "website_traffic": "+30%",
                "lead_generation": "+20%",
            },
        }

        return campaign


# Initialize AI systems
ai_systems = {
    "alex": AILeadGenerationSystem(),
    "maria": AIColdCallingSystem(),
    "david": AISalesOutreachSystem(),
    "emma": AISocialMediaSystem(),
    "carlos": AIEmailCampaignSystem(),
    "rachel": AIContentCreationSystem(),
    "mike": AISEOSystem(),
    "lisa": AIPPCSystem(),
    "tom": AIBrandStrategySystem(),
}


# AI Task Execution Functions
def execute_ai_task(ai_name, task_type, parameters):
    """Execute AI agent tasks with their specialized systems"""
    if ai_name not in ai_systems:
        return {"error": "AI agent not found"}

    ai_system = ai_systems[ai_name]

    try:
        if task_type == "find_prospects" and ai_name == "alex":
            return ai_system.find_prospects(
                parameters.get("industry"),
                parameters.get("company_size"),
                parameters.get("location"),
            )
        elif task_type == "generate_call_script" and ai_name == "maria":
            return ai_system.generate_call_script(
                parameters.get("prospect_data"), parameters.get("product_info")
            )
        elif task_type == "create_proposal" and ai_name == "david":
            return ai_system.create_proposal(
                parameters.get("prospect_data"), parameters.get("solution_details")
            )
        elif task_type == "create_content_calendar" and ai_name == "emma":
            return ai_system.create_content_calendar(
                parameters.get("themes"), parameters.get("frequency")
            )
        elif task_type == "create_email_campaign" and ai_name == "carlos":
            return ai_system.create_email_campaign(
                parameters.get("campaign_type"), parameters.get("target_audience")
            )
        elif task_type == "generate_blog_post" and ai_name == "rachel":
            return ai_system.generate_blog_post(
                parameters.get("topic"),
                parameters.get("target_keywords"),
                parameters.get("word_count", 800),
            )
        elif task_type == "analyze_website_performance" and ai_name == "mike":
            return ai_system.analyze_website_performance(parameters.get("domain"))
        elif task_type == "create_ppc_campaign" and ai_name == "lisa":
            return ai_system.create_ppc_campaign(
                parameters.get("campaign_type"),
                parameters.get("target_audience"),
                parameters.get("budget"),
            )
        elif task_type == "develop_brand_strategy" and ai_name == "tom":
            return ai_system.develop_brand_strategy(
                parameters.get("company_data"), parameters.get("target_audience")
            )
        else:
            return {"error": f"Task {task_type} not supported for {ai_name}"}

    except Exception as e:
        app.logger.error(f"Error executing AI task: {e}")
        return {"error": f"Task execution failed: {str(e)}"}


# AI Task API Endpoints
@app.route("/api/ai-tasks/<ai_name>/<task_type>", methods=["POST"])
@login_required
def execute_ai_task_api(ai_name, task_type):
    """API endpoint for executing AI agent tasks"""
    try:
        data = request.get_json()
        if not data:
            return jsonify(error="No parameters provided"), 400

        result = execute_ai_task(ai_name, task_type, data)

        if "error" in result:
            return jsonify(result), 400
        else:
            return jsonify(result), 200

    except Exception as e:
        app.logger.error(f"Error in AI task API: {e}")
        return jsonify(error="Internal server error"), 500


@app.route("/api/ai-systems/status", methods=["GET"])
@login_required
def get_ai_systems_status():
    """Get status of all AI systems"""
    try:
        status = {}
        for ai_name, system in ai_systems.items():
            status[ai_name] = {
                "system_type": system.__class__.__name__,
                "status": "active",
                "last_activity": datetime.utcnow().isoformat(),
                "capabilities": list(system.__class__.__dict__.keys()),
            }

        return jsonify(status), 200

    except Exception as e:
        app.logger.error(f"Error getting AI systems status: {e}")
        return jsonify(error="Failed to get AI systems status"), 500


if __name__ == "__main__":
    debug_mode = os.environ.get("FLASK_DEBUG", "0") == "1"
    app.run(debug=debug_mode, host="0.0.0.0")  # nosec B104
