import os
import sys
import json
import argparse
from datetime import datetime
import requests
from dotenv import load_dotenv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Load environment variables
load_dotenv()

# Get API key from environment variable or command line
API_KEY = os.getenv('GOOGLE_MAPS_API_KEY')

def get_coordinates(city, api_key):
    """Get coordinates for a city using Google Maps Geocoding API."""
    url = f"https://maps.googleapis.com/maps/api/geocode/json?address={city}&key={api_key}"
    response = requests.get(url)
    data = response.json()
    
    if data['status'] != 'OK':
        raise Exception(f"Geocoding failed: {data['status']}")
    
    location = data['results'][0]['geometry']['location']
    return location['lat'], location['lng']

def search_places(lat, lng, business_type, radius, api_key):
    """Search for places using Google Maps Places API."""
    url = f"https://maps.googleapis.com/maps/api/place/nearbysearch/json"
    params = {
        'location': f"{lat},{lng}",
        'radius': radius,
        'type': business_type,
        'key': api_key
    }
    
    response = requests.get(url, params=params)
    data = response.json()
    
    if data['status'] != 'OK':
        raise Exception(f"Places search failed: {data['status']}")
    
    return data['results']

def get_place_details(place_id, api_key):
    """Get detailed information about a place using Google Maps Places API."""
    url = f"https://maps.googleapis.com/maps/api/place/details/json"
    params = {
        'place_id': place_id,
        'fields': 'name,formatted_address,formatted_phone_number,website,rating,user_ratings_total,opening_hours,url,types,business_status,price_level,geometry',
        'key': api_key
    }
    
    response = requests.get(url, params=params)
    data = response.json()
    
    if data['status'] != 'OK':
        raise Exception(f"Place details failed: {data['status']}")
    
    return data['result']

def is_potential_lead(place):
    """Check if a place meets our criteria for a potential lead."""
    # Check if the place has fewer than 15 reviews
    if 'user_ratings_total' in place and place['user_ratings_total'] >= 15:
        return False
    
    # Check if the place has a website
    if 'website' in place:
        return False
    
    # Check if the place is operational
    if place.get('business_status') != 'OPERATIONAL':
        return False
    
    return True

def save_leads_to_excel(leads, city, business_type):
    """Save leads to a formatted Excel file."""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"leads_{city.replace(' ', '_')}_{business_type}_{timestamp}.xlsx"
    
    # Create a new workbook and select the active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Potential Leads"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Define headers
    headers = [
        'Business Name',
        'Address',
        'Phone Number',
        'Rating',
        'Number of Reviews',
        'Price Level',
        'Business Types',
        'Opening Hours',
        'Google Maps URL',
        'Latitude',
        'Longitude',
        'Business Status'
    ]
    
    # Write headers
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Write data
    for row, lead in enumerate(leads, 2):
        ws.cell(row=row, column=1, value=lead['name'])
        ws.cell(row=row, column=2, value=lead['address'])
        ws.cell(row=row, column=3, value=lead['phone'])
        ws.cell(row=row, column=4, value=lead['rating'])
        ws.cell(row=row, column=5, value=lead['reviews'])
        ws.cell(row=row, column=6, value=lead['price_level'])
        ws.cell(row=row, column=7, value=lead['business_types'])
        ws.cell(row=row, column=8, value=lead['opening_hours'])
        ws.cell(row=row, column=9, value=lead['google_maps_url'])
        ws.cell(row=row, column=10, value=lead['latitude'])
        ws.cell(row=row, column=11, value=lead['longitude'])
        ws.cell(row=row, column=12, value=lead['business_status'])
        
        # Apply borders and wrap text to all cells in the row
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 20
    
    # Save the workbook
    wb.save(filename)
    return filename

def main():
    parser = argparse.ArgumentParser(description='Find potential business leads using Google Maps API')
    parser.add_argument('city', help='City to search in (e.g., "New York, NY")')
    parser.add_argument('--business-type', default='plumber', help='Type of business to search for')
    parser.add_argument('--radius', type=int, default=5000, help='Search radius in meters')
    parser.add_argument('--api-key', help='Google Maps API key')
    parser.add_argument('--list-types', action='store_true', help='List available business types')
    
    args = parser.parse_args()
    
    # Use API key from command line if provided, otherwise use environment variable
    api_key = args.api_key or API_KEY
    if not api_key:
        print("Error: No API key provided. Please provide an API key using --api-key or set GOOGLE_MAPS_API_KEY environment variable.")
        sys.exit(1)
    
    if args.list_types:
        print("Available business types:")
        print("- plumber")
        print("- electrician")
        print("- restaurant")
        print("- cafe")
        print("- bar")
        print("- store")
        print("- hotel")
        print("- gym")
        print("- salon")
        print("- dentist")
        print("- doctor")
        print("- lawyer")
        print("- real_estate_agency")
        print("- car_dealer")
        print("- car_repair")
        print("- car_wash")
        print("- gas_station")
        print("- bank")
        print("- pharmacy")
        print("- supermarket")
        print("- shopping_mall")
        print("- movie_theater")
        print("- museum")
        print("- park")
        print("- school")
        print("- university")
        print("- hospital")
        print("- police")
        print("- fire_station")
        print("- post_office")
        print("- library")
        print("- church")
        print("- mosque")
        print("- synagogue")
        print("- temple")
        print("- cemetery")
        print("- zoo")
        print("- aquarium")
        print("- stadium")
        print("- airport")
        print("- train_station")
        print("- bus_station")
        print("- taxi_stand")
        print("- subway_station")
        print("- light_rail_station")
        print("- ferry_terminal")
        print("- transit_station")
        print("- point_of_interest")
        print("- establishment")
        sys.exit(0)
    
    try:
        print(f"\nSearching for {args.business_type}s in {args.city}...")
        lat, lng = get_coordinates(args.city, api_key)
        print(f"Coordinates: {lat}, {lng}")
        
        places = search_places(lat, lng, args.business_type, args.radius, api_key)
        print(f"Found {len(places)} places")
        
        leads = []
        for place in places:
            if is_potential_lead(place):
                details = get_place_details(place['place_id'], api_key)
                
                # Format opening hours
                opening_hours = "N/A"
                if 'opening_hours' in details and 'weekday_text' in details['opening_hours']:
                    opening_hours = "\n".join(details['opening_hours']['weekday_text'])
                
                # Format business types
                business_types = ", ".join(details.get('types', []))
                
                # Get coordinates
                latitude = details.get('geometry', {}).get('location', {}).get('lat', 'N/A')
                longitude = details.get('geometry', {}).get('location', {}).get('lng', 'N/A')
                
                lead = {
                    'name': details.get('name', 'N/A'),
                    'address': details.get('formatted_address', 'N/A'),
                    'phone': details.get('formatted_phone_number', 'N/A'),
                    'rating': details.get('rating', 'N/A'),
                    'reviews': details.get('user_ratings_total', 'N/A'),
                    'price_level': 'N/A' if 'price_level' not in details else '★' * details['price_level'],
                    'business_types': business_types,
                    'opening_hours': opening_hours,
                    'google_maps_url': details.get('url', 'N/A'),
                    'latitude': latitude,
                    'longitude': longitude,
                    'business_status': details.get('business_status', 'N/A')
                }
                leads.append(lead)
        
        if leads:
            filename = save_leads_to_excel(leads, args.city, args.business_type)
            print(f"\nFound {len(leads)} potential leads!")
            print(f"Results saved to {filename}")
        else:
            print("\nNo potential leads found.")
        
    except Exception as e:
        print(f"Error: {str(e)}")
        sys.exit(1)

if __name__ == '__main__':
    main() 